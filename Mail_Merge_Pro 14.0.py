import sys
import os
import json
import base64
import re
import socket
import openpyxl
from openpyxl.styles import PatternFill
import threading
import time
import requests

# --- PyQt5 Imports ---
try:
    from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                                 QHBoxLayout, QPushButton, QLabel, QListWidget, 
                                 QTextEdit, QMessageBox, QFileDialog, QInputDialog, 
                                 QCheckBox, QDialog, QFrame, QGridLayout, QGraphicsDropShadowEffect, 
                                 QSizePolicy, QProgressBar, QDialogButtonBox, QLineEdit, QTableWidget, 
                                 QTableWidgetItem, QHeaderView, QAbstractItemView, QAction, QMenu, QStackedLayout)
    from PyQt5.QtWebEngineWidgets import QWebEngineView
    from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QMutex, QWaitCondition, QSize, QPropertyAnimation, QRectF, QTimer, QRect
    from PyQt5.QtGui import QPixmap, QIcon, QFont, QColor, QPalette, QLinearGradient, QBrush, QGradient, QCursor, QTextCursor, QPainter, QPen
except ImportError:
    print("CRITICAL ERROR: PyQt5 or PyQtWebEngine is missing.")
    print("Please run: pip install PyQt5 PyQtWebEngine")
    sys.exit(1)

# --- Google API Imports ---
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders

# --- GLOBALS & CONSTANTS ---
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly',
          'https://www.googleapis.com/auth/gmail.send',
          'https://www.googleapis.com/auth/userinfo.profile']
PROGRESS_FILE = "mail_merge_progress.json"

# --- UTILITY FUNCTIONS ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def extract_body_and_attachments(payload, msg_id, service):
    html = ""
    attachments = []

    def walk(parts):
        nonlocal html
        for part in parts:
            mime = part.get('mimeType', '')
            headers = {h['name']: h['value'] for h in part.get('headers', [])}
            if mime == 'text/html' and 'body' in part and 'data' in part['body']:
                data = base64.urlsafe_b64decode(part['body']['data']).decode()
                html = data
            elif part.get('body', {}).get('attachmentId'):
                att = service.users().messages().attachments().get(
                    userId='me', messageId=msg_id, id=part['body']['attachmentId']).execute()
                file_data = base64.urlsafe_b64decode(att['data'])
                filename = part.get('filename', '')
                cid = headers.get('Content-ID')
                attachments.append((mime, filename, file_data, cid))
            if 'parts' in part:
                walk(part['parts'])

    if payload.get('parts'):
        walk(payload['parts'])
    elif 'body' in payload and 'data' in payload['body']:
        html = base64.urlsafe_b64decode(payload['body']['data']).decode()

    return html, attachments

def personalize(text, row_data, headers):
    for i, header in enumerate(headers):
        value = row_data[i]
        placeholder = f"{{{{{header}}}}}"
        
        if value is not None and str(value).strip() != "":
            text = text.replace(placeholder, str(value))
        else:
            esc_p = re.escape(placeholder)
            text = re.sub(r'^\s*' + esc_p + r'\s*$', '', text, flags=re.MULTILINE)
            text = re.sub(r',\s*' + esc_p, '', text)
            text = re.sub(esc_p + r'\s*,\s+', '', text)
            text = re.sub(r'\s+' + esc_p, '', text)
            text = re.sub(esc_p + r'\s+', '', text)
            text = re.sub(esc_p, '', text)
            
    text = re.sub(r'\s+([,.])', r'\1', text)
    text = re.sub(r'([,.])\1+', r'\1', text)
    text = re.sub(r'^\s*[,.]\s*', '', text)
    return text

def clean_personalization(text, row_data, headers):
    """
    Robust cleaning: Finds {{...}} blocks and strips internal HTML tags/whitespace
    before passing to the standard personalize function.
    """
    if not text: return ""

    # Pattern to find curly brace blocks, possibly containing HTML tags
    # We look for {{ (anything not }) }}
    # But specifically, we want to address {{<tags>var</tags>}} scenarios.
    
    def strip_tags(match):
        content = match.group(1) # The stuff inside {{...}}
        
        # 1. Remove all HTML tags
        clean_content = re.sub(r'<[^>]+>', '', content)
        
        # 2. Unescape common HTML entities if necessary
        clean_content = clean_content.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
        
        # 3. Optimize Whitespace (remove leading/trailing, single spaces)
        clean_content = clean_content.strip()
        
        return f"{{{{{clean_content}}}}}"

    # Regex: Matches {{...}} where ... is distinct from just }}
    # Using non-greedy match to find minimal pairs
    text = re.sub(r'\{\{(.+?)\}\}', strip_tags, text, flags=re.DOTALL)
    
    return personalize(text, row_data, headers)

def get_email_recipients(row_values, all_headers, cc_mode, global_cc, bcc_mode, global_bcc):
    """
    Resolves To, CC, and BCC for a given row.
    Returns (recipient, cc_string, bcc_string)
    """
    recipient = ""
    cc_str = ""
    bcc_str = ""
    
    # 1. Get Recipient (Email column)
    try:
        email_idx = -1
        # Find "Email" column case-insensitive
        for idx, h in enumerate(all_headers):
            if str(h).strip().lower() == "email":
                email_idx = idx
                break
        
        if email_idx != -1 and len(row_values) > email_idx:
            val = row_values[email_idx]
            if val:
                recipient = str(val).strip()
    except Exception:
        pass

    # 2. CC Logic
    if cc_mode == "global" and global_cc:
        raw_cc = global_cc
        emails = [e.strip() for e in re.split(r'[,\n\r]+', raw_cc) if e.strip()]
        cc_str = ", ".join(emails)
    elif cc_mode == "individual":
        try:
            cc_i = next(i for i, h in enumerate(all_headers) if str(h).lower() == "cc")
            if len(row_values) > cc_i and row_values[cc_i]: 
                cc_str = str(row_values[cc_i]).strip()
        except: pass

    # 3. BCC Logic
    if bcc_mode == "global" and global_bcc:
        raw_bcc = global_bcc
        emails = [e.strip() for e in re.split(r'[,\n\r]+', raw_bcc) if e.strip()]
        bcc_str = ", ".join(emails)
    elif bcc_mode == "individual":
        try:
            bcc_i = next(i for i, h in enumerate(all_headers) if str(h).lower() == "bcc")
            if len(row_values) > bcc_i and row_values[bcc_i]:
                bcc_str = str(row_values[bcc_i]).strip()
        except: pass
        
    return recipient, cc_str, bcc_str

# --- WORKER THREAD FOR SENDING EMAILS ---
class EmailWorker(QThread):
    log_signal = pyqtSignal(str, str) # msg, color
    progress_signal = pyqtSignal(int)
    # Updated Signal to include CC string
    # Updated Signal to include CC and BCC string
    preview_signal = pyqtSignal(str, str, str, str, str) # subject, body, recipient, cc, bcc
    live_preview_signal = pyqtSignal(int, list, str) # row_index, row_values, status
    finished_signal = pyqtSignal(int, int) # sent, failed
    stopped_signal = pyqtSignal(int, int, int) # sent_session, failed_session, pending_total
    error_signal = pyqtSignal(str)

    def __init__(self, service, excel_path, draft_id, start_row, cc_mode, global_cc, bcc_mode, global_bcc, display_name, user_email, total_rows=None, is_resume=False, attachment_mode=True, attachment_empty_rule="yes"):
        super().__init__()
        self.service = service
        self.excel_path = excel_path
        self.draft_id = draft_id
        self.start_row = start_row
        self.cc_mode = cc_mode
        self.global_cc = global_cc
        self.bcc_mode = bcc_mode
        self.global_bcc = global_bcc
        self.display_name = display_name
        self.user_email = user_email
        self.total_rows = total_rows
        self.is_resume = is_resume
        self.attachment_mode = attachment_mode # True = Send All, False = Conditional
        self.attachment_empty_rule = attachment_empty_rule # "yes" or "no" for empty cells in conditional mode
        
        self.is_running = True


    def run(self):
        sent_count = 0
        fail_count = 0

        try:
            # Load Draft Data
            draft_detail = self.service.users().drafts().get(userId='me', id=self.draft_id).execute()
            msg0 = draft_detail['message']
            payload = msg0['payload']
            subject_tmpl = next((h['value'] for h in payload.get('headers', []) if h['name'] == 'Subject'), '(No Subject)')
            body_html_tmpl, attachments = extract_body_and_attachments(payload, msg0['id'], self.service)

            # Load Excel
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            # Headers & Indexing
            headers = []
            visible_indexes = []
            for idx, cell in enumerate(ws[1]):
                if not ws.column_dimensions[cell.column_letter].hidden:
                    headers.append(cell.value)
                    visible_indexes.append(idx)
            
            all_headers = [c.value for c in ws[1]]
            email_idx = all_headers.index("Email")
            
            # Status Column Logic (3-Column System)
            headers_lower = [str(h).strip().lower() for h in all_headers]
            
            def get_col_idx(names):
                for name in names:
                    if name in headers_lower: return headers_lower.index(name)
                return -1

            col_status = get_col_idx(['status', 'start'])
            col_stop = get_col_idx(['stop', 'stopped'])
            col_resume = get_col_idx(['resume', 'resumed'])
            
            # Attachment Control Column
            col_attachments = get_col_idx(['attachment', 'attachments', 'send attachment', 'send attachments', 'include attachments'])
            
            # Ensure Status/Stop/Resume columns exist
            if col_status == -1:
                 ws.cell(row=1, column=len(all_headers)+1).value = "Status"
                 all_headers.append("Status")
                 col_status = len(all_headers) - 1
                 headers_lower.append('status')
            
            if col_stop == -1:
                 ws.cell(row=1, column=len(all_headers)+1).value = "Stop"
                 all_headers.append("Stop")
                 col_stop = len(all_headers) - 1
                 headers_lower.append('stop')
                 
            if col_resume == -1:
                 ws.cell(row=1, column=len(all_headers)+1).value = "Resume"
                 all_headers.append("Resume")
                 col_resume = len(all_headers) - 1
                 headers_lower.append('resume')

            self.log_signal.emit(f"🚀 Starting from Row {self.start_row}...", "#17A2B8")

            # Calculate Total Rows for Progress Bar
            max_row = ws.max_row
            # We use self.total_rows passed from outside for LOGGING consistency, 
            # but for progress bar PERCENTAGE we still use relative progress if desired, 
            # OR we can switch progress bar to be absolute.
            # Let's keep progress bar relative to "this run" but logs absolute "current/total".
            
            total_to_process = max_row - self.start_row + 1
            if total_to_process < 1: total_to_process = 1
            
            # If total_rows not provided (e.g. Resume), estimate using Email column
            if not self.total_rows:
                # Find Email Column
                email_idx = -1
                _all_headers = [c.value for c in ws[1]]
                for i, h in enumerate(_all_headers):
                     if str(h).strip().lower() == 'email':
                         email_idx = i
                         break
                
                if email_idx != -1:
                    # Count non-empty emails
                    count = 0
                    try:
                        for row in ws.iter_rows(min_row=2, min_col=email_idx+1, max_col=email_idx+1):
                            if row and len(row) > 0 and row[0].value:
                                count += 1
                    except Exception as e:
                        self.log_signal.emit(f"⚠️ Debug: Count Error {e}", "#FFC107")
                    self.total_rows = count
                else:
                    # Fallback
                    self.total_rows = max_row - 1
                
                if self.total_rows < 1: self.total_rows = 1


            # Iterate Rows
            processed_count = 0
            for idx, row in enumerate(ws.iter_rows(min_row=self.start_row), start=self.start_row):
                try:
                    if not self.is_running:
                        self.save_progress_and_stop(idx, wb, sent_count, fail_count)
                        wb.save(self.excel_path) # CRITICAL FIX: Save on Stop
                        return
                        
                    # Debug Point 1: Row Loaded
                    # self.log_signal.emit(f"Debug: Row {idx} raw len {len(row)}", "#17A2B8")

                    processed_count += 1
                    
                    # Update Progress Bar
                    if self.total_rows and self.total_rows > 0:
                         progress_percent = int(((idx - 1) / self.total_rows) * 100)
                    else:
                         progress_percent = int((processed_count / total_to_process) * 100)
                    
                    self.progress_signal.emit(progress_percent)

                    if not row: continue # Empty row tuple
                    
                    # Safe Email Access
                    try:
                         if len(row) > email_idx and row[email_idx].value:
                             recipient = row[email_idx].value
                         else:
                             continue # No email
                    except IndexError:
                         continue

                    row_values = [cell.value for cell in row]
                    
                    # Safety Pad: Ensure row_values matches expected header length
                    if len(row_values) < len(all_headers):
                         row_values.extend([None] * (len(all_headers) - len(row_values)))
                         
                    # Debug Point 2: Padding Done
                    
                    filtered_row = [row_values[i] for i in visible_indexes]

                    # Emit "Sending..." status
                    self.live_preview_signal.emit(idx, row_values, "Sending...")

                    # Personalize (Potential Crash Point)
                    subj_p = personalize(subject_tmpl, filtered_row, headers)
                    body_p = personalize(body_html_tmpl, filtered_row, headers)

                    # --- DETERMINE CC & BCC ---
                    current_cc = ""
                    current_bcc = ""
                    
                    # CC Logic
                    if self.cc_mode == "global" and self.global_cc:
                        raw_cc = self.global_cc
                        emails = [e.strip() for e in re.split(r'[,\n\r]+', raw_cc) if e.strip()]
                        current_cc = ", ".join(emails)
                    elif self.cc_mode == "individual":
                        try:
                            # Case-insensitive search for CC
                            cc_i = next(i for i, h in enumerate(all_headers) if str(h).lower() == "cc")
                            if len(row_values) > cc_i and row_values[cc_i]: 
                                current_cc = str(row_values[cc_i]).strip()
                        except: pass

                    # BCC Logic
                    if self.bcc_mode == "global" and self.global_bcc:
                        raw_bcc = self.global_bcc
                        emails = [e.strip() for e in re.split(r'[,\n\r]+', raw_bcc) if e.strip()]
                        current_bcc = ", ".join(emails)
                    elif self.bcc_mode == "individual":
                        try:
                            # Case-insensitive search for BCC
                            bcc_i = next(i for i, h in enumerate(all_headers) if str(h).lower() == "bcc")
                            if len(row_values) > bcc_i and row_values[bcc_i]:
                                current_bcc = str(row_values[bcc_i]).strip()
                        except: pass



                    # --- SENDING LOGIC ---
                    msg = MIMEMultipart('related')
                    msg['From'] = f"{self.display_name} <{self.user_email}>"
                    msg['To'] = recipient
                    msg['Subject'] = subj_p

                    # Apply CC
                    if current_cc:
                        msg['Cc'] = current_cc
                    
                    # Apply BCC
                    if current_bcc:
                        msg['Bcc'] = current_bcc

                    alt = MIMEMultipart('alternative')
                    alt.attach(MIMEText(body_p, 'html'))
                    msg.attach(alt)

                    # Attachments Logic
                    # Determine if we should send attachments for this user
                    send_attachments_for_user = True
                    
                    if not self.attachment_mode: # Conditional Mode
                        # Check Excel Column
                        if col_attachments != -1:
                            val = row_values[col_attachments]
                            str_val = str(val).strip().lower() if val else ""
                            
                            if not str_val: # Empty
                                 if self.attachment_empty_rule == "no":
                                     send_attachments_for_user = False
                            elif str_val in ['no', 'n', 'false', '0']:
                                send_attachments_for_user = False
                        else:
                            # CRITICAL SAFETY: If conditional mode but column missing, DO NOT SEND.
                            send_attachments_for_user = False
                            if idx == self.start_row: # Log once
                                self.log_signal.emit("⚠️ Formatting Error: 'Send Attachments' column not found. Skipping attachments.", "#FFC107")
                                
                                 
                    if send_attachments_for_user:
                        for mime, fname, fdata, cid in attachments:
                            if mime.startswith('image/') and cid:
                                img = MIMEImage(fdata, _subtype=mime.split('/')[1])
                                img.add_header('Content-ID', cid)
                                img.add_header('Content-Disposition', 'inline', filename=fname)
                                msg.attach(img)
                            else:
                                part = MIMEBase(*mime.split('/', 1))
                                part.set_payload(fdata)
                                encoders.encode_base64(part)
                                part.add_header('Content-Disposition', 'attachment', filename=fname)
                                msg.attach(part)


                    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
                    self.service.users().messages().send(userId='me', body={'raw': raw}).execute()
                    
                    status_msg = "Sent"
                    if attachments:
                        status_msg = "Sent with Attachment" if send_attachments_for_user else "Sent without Attachment"
                    
                    log_msg = f"[{idx - 1}/{self.total_rows}] ✅ {status_msg} to {recipient}"
                    self.log_signal.emit(log_msg, "#28A745")
                    
                    
                    # --- 3-Column Logic ---
                    
                    # 1. Update "Status" Column
                    # 1. Update "Status" Column
                    if col_status != -1:
                        cell = ws.cell(row=idx, column=col_status + 1)
                        cell.value = status_msg
                        
                        # Color Logic
                        if "without Attachment" in status_msg:
                             # Light Green for "Sent without Attachment"
                             cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                             cell.font = openpyxl.styles.Font(color="006100") 
                        else:
                             # Dark Green for "Sent with Attachment"
                             cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
                             cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

                    # 2. Update "Resume" Column (Yellow "Resumed")
                    # Only for the FIRST processed row if this is a Resume session
                    if self.is_resume and processed_count == 1:
                        if col_resume != -1:
                            cell = ws.cell(row=idx, column=col_resume + 1)
                            cell.value = "Resumed"
                            cell.fill = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid") # Yellow
                        self.live_preview_signal.emit(idx, row_values, "Resumed")
                    else:
                        self.live_preview_signal.emit(idx, row_values, "Sent")

                    sent_count += 1
                
                except IndexError as ie:
                    # Capture exact list error
                    import traceback
                    tb = traceback.format_exc()
                    self.log_signal.emit(f"❌ Index Error Row {idx}: {ie}\nTraceback:\n{tb}", "#DC3545")
                    fail_count += 1
                    
                except Exception as e:
                    import traceback
                    tb = traceback.format_exc()
                    self.log_signal.emit(f"❌ Failed to {recipient if 'recipient' in locals() else 'Unknown'}: {e}\nTraceback:\n{tb}", "#DC3545")
                    
                    # Error in "Status" column? Or Stop? usually Status.
                    if 'col_status' in locals() and col_status != -1:
                        cell = ws.cell(row=idx, column=col_status + 1)
                        cell.value = f"Error: {str(e)}"
                        cell.fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid") # Red

                    if 'row_values' in locals():
                        self.live_preview_signal.emit(idx, row_values, "Error")
                    fail_count += 1

            # Done
            wb.save(self.excel_path)
            if os.path.exists(PROGRESS_FILE): os.remove(PROGRESS_FILE)
            self.finished_signal.emit(sent_count, fail_count)

        except Exception as e:
            self.error_signal.emit(f"Critical Worker Error: {e}")

    def save_progress_and_stop(self, idx, wb, sent_count, fail_count):
        # Mark current row as Stopped if not sent
        try:
           ws = wb.active
           headers = [str(c.value).strip().lower() for c in ws[1]]
           
           # Find "Stop" column
           col_stop = -1
           for i, h in enumerate(headers):
               if h in ['stop', 'stopped']:
                   col_stop = i
                   break
           
           if col_stop != -1:
               cell = ws.cell(row=idx, column=col_stop + 1)
               cell.value = "Stopped"
               cell.fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid") # Red
        except: pass
        
        wb.save(self.excel_path)
        with open(PROGRESS_FILE, 'w') as f:
            json.dump({"last_row": idx}, f)
        # Log exactly where we are saving, so the user knows where Resume will start
        self.log_signal.emit(f"💾 Progress saved. Resume will start from Email #{idx - 1}.", "#FD7E14")
        
        # Calculate Pending
        pending = 0
        if self.total_rows:
            pending = max(0, self.total_rows - (idx - 2))

        self.stopped_signal.emit(sent_count, fail_count, pending)
        self.finished_signal.emit(-1, -1) # -1 indicates stopped

    def stop(self):
        self.is_running = False




# --- CUSTOM RESIZABLE INPUT DIALOG ---
class ResizableInputDialog(QDialog):
    def __init__(self, title, label_text, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(400, 250) # Set a larger default size
        
        # Style (Matches App)
        self.setStyleSheet("""
            QDialog { background-color: #F8F9FA; }
            QLabel { color: #333; font-weight: bold; font-size: 14px; }
            QTextEdit { 
                background-color: white; border: 1px solid #DEE2E6; 
                border-radius: 6px; padding: 5px; font-size: 13px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(label_text))
        
        self.text_input = QTextEdit()
        self.text_input.setPlaceholderText("Enter email(s) here...")
        layout.addWidget(self.text_input)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def get_text(self):
        return self.text_input.toPlainText().strip()

class ModernAttachmentDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.resize(550, 450)
        self.result_rule = "yes" # Default
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        container = QFrame()
        container.setStyleSheet("background-color: white; border-radius: 12px; border: 1px solid #E9ECEF;")
        # Dropshadow
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0,0,0,50))
        container.setGraphicsEffect(shadow)
        layout.addWidget(container)
        
        self.inner_layout = QVBoxLayout(container)
        self.inner_layout.setContentsMargins(25, 25, 25, 25)
        self.inner_layout.setSpacing(15)
        
        # --- PHASE 1: WARNING ---
        self.icon_lbl = QLabel("📎")
        self.icon_lbl.setAlignment(Qt.AlignCenter)
        self.icon_lbl.setStyleSheet("font-size: 48px; margin-bottom: 10px;")
        
        self.title_lbl = QLabel("Enable Conditional Attachments?")
        self.title_lbl.setAlignment(Qt.AlignCenter)
        self.title_lbl.setStyleSheet("font-size: 20px; font-weight: bold; color: #343A40;")
        
        self.msg_lbl = QLabel(
            "If enabled, we will scan your Excel file for a <b>'Send Attachments'</b> column.<br><br>"
            "• <b>Yes</b>: Attachments sent.<br>"
            "• <b>No</b>: Attachments removed.<br><br>"
            "Are you sure you want to proceed?"
        )
        self.msg_lbl.setWordWrap(True)
        self.msg_lbl.setStyleSheet("font-size: 14px; color: #495057; line-height: 1.5;")
        
        # Determine Phase 1 Buttons
        self.btn_layout = QHBoxLayout()
        
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setCursor(Qt.PointingHandCursor)
        self.btn_cancel.setStyleSheet("""
            QPushButton { background-color: #E9ECEF; color: #333; border-radius: 6px; padding: 10px; font-weight: bold; border: 1px solid #CED4DA; }
            QPushButton:hover { background-color: #DEE2E6; }
        """)
        self.btn_cancel.clicked.connect(self.reject)
        
        self.btn_yes = QPushButton("Enable Conditional Mode")
        self.btn_yes.setCursor(Qt.PointingHandCursor)
        self.btn_yes.setStyleSheet("""
            QPushButton { background-color: #0D6EFD; color: white; border-radius: 6px; padding: 10px; font-weight: bold; }
            QPushButton:hover { background-color: #0B5ED7; }
        """)
        self.btn_yes.clicked.connect(self.run_validation)
        
        self.btn_layout.addWidget(self.btn_cancel)
        self.btn_layout.addWidget(self.btn_yes)
        
        self.inner_layout.addWidget(self.icon_lbl)
        self.inner_layout.addWidget(self.title_lbl)
        self.inner_layout.addWidget(self.msg_lbl)
        self.inner_layout.addStretch()
        self.inner_layout.addLayout(self.btn_layout)

        # Reference for Phase 2 widgets
        self.table = None
    
    def run_validation(self):
        # Trigger parent to scan
        parent = self.parent()
        if hasattr(parent, 'scan_excel_for_empty_attachments'):
            empty_rows = parent.scan_excel_for_empty_attachments()
            
            if empty_rows:
                self.show_phase_2(empty_rows)
            else:
                self.accept() # All good, no empty values
        else:
            self.accept()

    def show_phase_2(self, empty_rows):
        # Clear Layout
        for i in reversed(range(self.inner_layout.count())): 
            w = self.inner_layout.itemAt(i).widget()
            l = self.inner_layout.itemAt(i).layout()
            if w: w.setParent(None)
            if l: 
                while l.count(): 
                    item = l.takeAt(0)
                    if item.widget(): item.widget().setParent(None)
                l.setParent(None)

        self.resize(600, 550)
        
        # New Content -- NON-CODER FRIENDLY
        lbl_warn = QLabel("⚠️ Action Needed: Unclear Instructions")
        lbl_warn.setStyleSheet("font-size: 18px; font-weight: bold; color: #DC3545;")
        lbl_warn.setAlignment(Qt.AlignCenter)
        self.inner_layout.addWidget(lbl_warn)
        
        lbl_info = QLabel(
            f"We found <b>{len(empty_rows)} users</b> where the <b>'Send Attachments'</b> cell is empty.<br>"
            "Please update your Excel file manually. Enter <b>'Yes'</b> or <b>'No'</b> for these users."
        )
        lbl_info.setStyleSheet("color: #333; font-size: 14px; margin-bottom: 10px;")
        lbl_info.setWordWrap(True)
        self.inner_layout.addWidget(lbl_info)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(3) # Added Status Column
        self.table.setHorizontalHeaderLabels(["Name", "Email", "Send Attachments"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.setRowCount(len(empty_rows))
        self.table.setStyleSheet("border: 1px solid #DEE2E6; border-radius: 4px;")
        
        for r, row_data in enumerate(empty_rows):
            self.table.setItem(r, 0, QTableWidgetItem(str(row_data[0])))
            self.table.setItem(r, 1, QTableWidgetItem(str(row_data[1])))
            
            # Status Item (Empty)
            item_status = QTableWidgetItem("(Blank)")
            item_status.setForeground(QColor("#DC3545")) # Red text to highlight error
            item_status.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 2, item_status)
            
        self.inner_layout.addWidget(self.table)
        
        # Action Buttons
        actions_layout = QHBoxLayout()
        actions_layout.setSpacing(10)
        actions_layout.addStretch()
        
        btn_close = QPushButton("Close")
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.setStyleSheet("""
            QPushButton { background-color: #6C757D; color: white; padding: 10px 30px; border-radius: 6px; border: none; font-weight: bold;}
            QPushButton:hover { background-color: #5C636A; }
        """)
        # Since we are removing choice, clicking close basically means "Cancel" / Reject
        # The parent logic will catch rejection and stop the process.
        btn_close.clicked.connect(self.reject) 
        
        actions_layout.addWidget(btn_close)
        actions_layout.addStretch()
        
        self.inner_layout.addLayout(actions_layout)
        
    def finalize(self, rule):
        self.result_rule = rule
        self.accept()

# --- WORKER TO LOAD DATA BEFORE PREVIEW ---
class DataLoadingWorker(QThread):
    data_loaded = pyqtSignal(object, object, list, list, list) # draft_data, wb, all_headers, visible_headers, rows
    status_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, service, draft_id, excel_path):
        super().__init__()
        self.service = service
        self.draft_id = draft_id
        self.excel_path = excel_path

    def run(self):
        try:
            self.status_signal.emit("Downloading Draft...")
            # 1. Load Draft
            draft_detail = self.service.users().drafts().get(userId='me', id=self.draft_id).execute()
            msg0 = draft_detail['message']
            payload = msg0['payload']
            subject_tmpl = next((h['value'] for h in payload.get('headers', []) if h['name'] == 'Subject'), '(No Subject)')
            body_html_tmpl, attachments = extract_body_and_attachments(payload, msg0['id'], self.service)
            
            draft_data = {
                'id': self.draft_id, # Added ID to fix KeyError
                'subject': subject_tmpl,
                'body': body_html_tmpl,
                'attachments': attachments
            }

            self.status_signal.emit("Reading Excel...")
            # 2. Load Excel (Optimized)
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            ws = wb.active
            
            headers = []
            visible_indexes = []
            all_headers = []
            
            # Read Headers
            # ws[1] in read_only mode returns cell objects
            row1 = next(ws.iter_rows(min_row=1, max_row=1))
            
            for idx, cell in enumerate(row1):
                val = cell.value
                all_headers.append(val)
                # In read_only, column_dimensions might not be available or accurate for 'hidden'
                # fallback to showing all if specific hidden check fails or is complex
                # For now, let's assume all visible in read_only or check if we can skip logic
                # To be safe and fast, we'll treat all as visible or just check value
                if val:
                    headers.append(val)
                    visible_indexes.append(idx)
            
            # Find Email Index
            email_idx = -1
            for i, h in enumerate(all_headers):
                if str(h).strip().lower() == "email":
                    email_idx = i
                    break

            # --- FORCE STATUS COLUMNS ---
            extra_headers = ["Status", "Stop", "Resume"]
            missing_headers = []
            lower_headers = [str(h).lower() for h in all_headers]
            
            for h in extra_headers:
                if h.lower() not in lower_headers:
                    missing_headers.append(h)
            
            # Add to headers (so they appear in preview)
            all_headers.extend(missing_headers)
            headers.extend(missing_headers) # Assume we want to see them
            # -----------------------------

            self.status_signal.emit("Processing rows...")
            # Read All Rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                # row is tuple of cells
                row_values = [c.value for c in row]
                
                # Check for Email
                if email_idx != -1 and len(row_values) > email_idx:
                    if not row_values[email_idx]:
                        continue # Skip empty email
                else:
                    # If we can't find email column or value, safe to skip or keep? 
                    # If strictly following logic, if no email, we can't send.
                    continue

                # Filter for visible only (for personalization)
                filtered_row = [row_values[i] for i in visible_indexes if i < len(row_values)]
                
                # --- PAD WITH EMPTY STRINGS FOR MISSING HEADERS ---
                # This ensures personalization doesn't crash and columns show up empty
                if missing_headers:
                    pad = [""] * len(missing_headers)
                    filtered_row.extend(pad)
                    # We don't necessarily need to pad row_values unless used elsewhere by index
                    # But keeping simple.
                # --------------------------------------------------

                 # Store tuple: (original_row_obj, row_values, filtered_values, row_index_1_based)
                 # Note: in read_only, row[0].row exists.
                rows.append({
                    'values': row_values,
                    'filtered': filtered_row,
                    'index': row[0].row if row else 0
                })
            
            wb.close()
            self.data_loaded.emit(draft_data, None, all_headers, headers, rows)
            
        except Exception as e:
            self.error_signal.emit(str(e))

# --- ADVANCED PREVIEW DIALOG ---
class AdvancedPreviewDialog(QDialog):
    start_sending = pyqtSignal()

    def __init__(self, parent, draft_data, all_headers, visible_headers, rows, cc_mode, global_cc, bcc_mode, global_bcc, attachment_mode=True):
        super().__init__(parent)
        self.draft_data = draft_data
        self.all_headers = all_headers
        self.visible_headers = visible_headers
        self.rows = rows
        self.cc_mode = cc_mode
        self.global_cc = global_cc
        self.bcc_mode = bcc_mode
        self.global_bcc = global_bcc
        self.attachment_mode = attachment_mode # NEW
        
        self.current_idx = 0
        self.total = len(self.rows)
        
        self.setWindowTitle("Preparing to Send - Preview")
        self.resize(900, 750) # Reverted to original spacious size
        # Add Maximize Button
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
        self.setStyleSheet("""
            QDialog { background-color: #F8F9FA; }
            QLabel { color: #333; font-size: 14px; }
            QPushButton { padding: 8px 15px; border-radius: 6px; font-weight: bold; }
        """)
        
        self.init_ui()
        self.load_preview(0)

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Header Info Card
        info_card = QFrame()
        info_card.setStyleSheet("background-color: white; border: 1px solid #DEE2E6; border-radius: 8px; padding: 10px;")
        info_layout = QGridLayout(info_card)
        
        self.lbl_idx = QLabel()
        self.lbl_idx.setStyleSheet("color: #6C757D; font-weight: bold;")
        info_layout.addWidget(self.lbl_idx, 0, 0, 1, 2)
        
        info_layout.addWidget(QLabel("<b>To:</b>"), 1, 0)
        self.lbl_to = QLabel()
        self.lbl_to.setStyleSheet("color: #0D6EFD; font-weight: bold;")
        info_layout.addWidget(self.lbl_to, 1, 1)
        
        info_layout.addWidget(QLabel("<b>CC:</b>"), 2, 0)
        self.lbl_cc = QLabel()
        info_layout.addWidget(self.lbl_cc, 2, 1)

        info_layout.addWidget(QLabel("<b>BCC:</b>"), 3, 0)
        self.lbl_bcc = QLabel()
        info_layout.addWidget(self.lbl_bcc, 3, 1)
        
        info_layout.addWidget(QLabel("<b>Subject:</b>"), 4, 0)
        self.lbl_subj = QLabel()
        info_layout.addWidget(self.lbl_subj, 4, 1)

        info_layout.addWidget(QLabel("<b>Attachments:</b>"), 5, 0)
        self.lbl_att_status = QLabel()
        info_layout.addWidget(self.lbl_att_status, 5, 1)
        
        layout.addWidget(info_card)
        
        # Browser
        self.browser = QWebEngineView()
        self.browser.setStyleSheet("border: 1px solid #DEE2E6;")
        layout.addWidget(self.browser)
        
        # Navigation Bar
        nav_layout = QHBoxLayout()
        nav_layout.addStretch() # Center alignment
        
        self.btn_prev = QPushButton("<< Previous")
        self.btn_prev.setCursor(Qt.PointingHandCursor)
        self.btn_prev.setStyleSheet("""
            QPushButton {
                background-color: #6C757D;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #5A6268; }
            QPushButton:disabled { background-color: #E9ECEF; color: #ADB5BD; }
        """)
        self.btn_prev.clicked.connect(self.prev_mail)
        nav_layout.addWidget(self.btn_prev)
        
        # Add small spacing around counter
        nav_layout.addSpacing(15)
        
        self.lbl_counter = QLabel()
        self.lbl_counter.setAlignment(Qt.AlignCenter)
        self.lbl_counter.setStyleSheet("font-size: 14px; font-weight: bold; color: #333;")
        nav_layout.addWidget(self.lbl_counter)
        
        nav_layout.addSpacing(15)
        
        self.btn_next = QPushButton("Next >>")
        self.btn_next.setCursor(Qt.PointingHandCursor)
        self.btn_next.setStyleSheet("""
            QPushButton {
                background-color: #0D6EFD;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #0B5ED7; }
            QPushButton:disabled { background-color: #E9ECEF; color: #ADB5BD; }
        """)
        self.btn_next.clicked.connect(self.next_mail)
        nav_layout.addWidget(self.btn_next)
        
        nav_layout.addStretch() # Center alignment
        
        layout.addLayout(nav_layout)
        
        # Add spacing between Nav and Footer
        layout.addSpacing(20)
        
        # Action Bar
        action_layout = QHBoxLayout()
        
        # 1. Confirmation Checkbox (Left Side)
        self.chk_confirm = QCheckBox("I confirm I want to start sending")
        self.chk_confirm.setStyleSheet("font-weight: bold; color: #495057; margin-right: 15px; margin-left: 5px;")
        self.chk_confirm.stateChanged.connect(self.on_confirm_toggled)
        action_layout.addWidget(self.chk_confirm)
        
        # Spacer to push buttons to right
        action_layout.addStretch()
        
        # 2. Buttons (Right Side)
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setCursor(Qt.PointingHandCursor)
        self.btn_cancel.setFixedHeight(40) # Fixed Height
        self.btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #DC3545;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 15px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover { background-color: #C82333; }
        """)
        self.btn_cancel.clicked.connect(self.reject)
        action_layout.addWidget(self.btn_cancel)
        
        action_layout.addSpacing(10) # Gap between buttons
        
        self.btn_start = QPushButton("🚀 Start Sending")
        self.btn_start.setCursor(Qt.PointingHandCursor)
        self.btn_start.setEnabled(False) 
        self.btn_start.setFixedHeight(40) # Fixed Height
        self.btn_start.setStyleSheet("""
            QPushButton {
                background-color: #28A745;
                color: white;
                border: none;
                font-size: 15px;
                border-radius: 6px;
                font-weight: bold;
                min-width: 160px;
                padding-left: 20px;
                padding-right: 20px;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #E9ECEF; color: #ADB5BD; }
        """)
        self.btn_start.clicked.connect(self.on_start)
        action_layout.addWidget(self.btn_start)
        
        layout.addLayout(action_layout)

    def on_confirm_toggled(self, state):
        self.btn_start.setEnabled(state == Qt.Checked)

    def load_preview(self, idx):
        if not self.rows:
            self.lbl_to.setText("(No Data)")
            return

        self.current_idx = idx
        row_data = self.rows[idx]
        
        # Safety Pad (Same as Worker)
        # Ensure values list is long enough to cover all headers (including Send Attachments at end)
        if len(row_data['values']) < len(self.all_headers):
             row_data['values'].extend([None] * (len(self.all_headers) - len(row_data['values'])))

        # Resolve Recipients
        recip, cc, bcc = get_email_recipients(
            row_data['values'], self.all_headers, 
            self.cc_mode, self.global_cc, 
            self.bcc_mode, self.global_bcc
        )
        
        self.lbl_idx.setText(f"Previewing Email #{row_data['index'] - 1}")
        self.lbl_to.setText(recip)
        self.lbl_cc.setText(cc)
        self.lbl_bcc.setText(bcc)
        
        # Personalize
        # We need visible headers for clean_personalization
        # self.visible_headers should be passed in init, or we use all_headers if strictly matching logic
        # Assuming visible_headers logic is correct
        subj_p = clean_personalization(self.draft_data['subject'], row_data['filtered'], self.visible_headers)
        body_p = clean_personalization(self.draft_data['body'], row_data['filtered'], self.visible_headers)
        
        
        self.lbl_subj.setText(subj_p)
        self.browser.setHtml(body_p)
        
        # Attachment Status
        if self.attachment_mode:
            # show actual attachments
            att_names = [a[1] for a in self.draft_data.get('attachments', [])]
            if att_names:
                 names_str = ", ".join(att_names)
                 self.lbl_att_status.setText(f"✅ Yes ({names_str})")
            else:
                 self.lbl_att_status.setText("✅ Yes (No files in draft)")
            self.lbl_att_status.setStyleSheet("color: #198754; font-weight: bold;")
        else:
             # Check Excel Logic
             # Need headers to find column
             col_att = -1
             headers_lower = [str(h).strip().lower() for h in self.all_headers]
             
             # DEBUG: Print headers to see what we are working with
             print(f"DEBUG PREVIEW: Headers: {headers_lower}")
             
             for name in ['attachment', 'attachments', 'send attachments', 'send attachment']:
                  if name in headers_lower: 
                       col_att = headers_lower.index(name)
                       print(f"DEBUG PREVIEW: Found '{name}' at index {col_att}")
                       break
             
             # Need filenames too
             att_names = [a[1] for a in self.draft_data.get('attachments', [])]
             names_str = ", ".join(att_names) if att_names else "No files"
             
             status_text = "⚠️ Unresolved (Empty)"
             status_color = "#FD7E14" # Orange
             
             if col_att != -1:
                 # Check value
                 if len(row_data['values']) > col_att:
                     val = row_data['values'][col_att]
                     str_val = str(val).strip().lower() if val else ""
                     
                     if str_val in ['no', 'n', 'false', '0']:
                          status_text = "❌ No (Skipped)"
                          status_color = "#DC3545" # Red
                     elif str_val: # Yes or other value
                          status_text = f"✅ Yes ({names_str})"
                          status_color = "#198754" # Green
                     else:
                          # Empty value found
                          status_text = "⚠️ Unresolved (Cell Empty)"
                          status_color = "#FD7E14"
                 else:
                     status_text = "⚠️ Error (Row Short)"
                     status_color = "#DC3545"
             else:
                 # Column not found
                 status_text = "⚠️ Error (Column Not Found)"
                 status_color = "#DC3545"
             
             self.lbl_att_status.setText(status_text)
             self.lbl_att_status.setStyleSheet(f"color: {status_color}; font-weight: bold;")
        
        self.lbl_counter.setText(f"Email {idx + 1} of {self.total}")
        
        self.btn_prev.setEnabled(idx > 0)
        self.btn_next.setEnabled(idx < self.total - 1)

    def prev_mail(self):
        if self.current_idx > 0:
            self.load_preview(self.current_idx - 1)

    def next_mail(self):
        if self.current_idx < self.total - 1:
            self.load_preview(self.current_idx + 1)
            
    def set_visible_headers(self, headers):
        self.visible_headers = headers

    def on_start(self):
        self.start_sending.emit()
        self.accept()


# --- CUSTOM UI DIALOGS ---
class ModernInfoDialog(QDialog):
    def __init__(self, parent, title, message_html, icon_text="ℹ️", accent_color="#0D6EFD"):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.resize(380, 250)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        self.container = QFrame()
        self.container.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 12px; border: 1px solid #E9ECEF; }
        """)
        
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 30))
        self.container.setGraphicsEffect(shadow)
        layout.addWidget(self.container)
        
        content = QVBoxLayout(self.container)
        content.setContentsMargins(20, 20, 20, 20)
        content.setSpacing(15)
        
        # Icon & Title
        header = QHBoxLayout()
        lbl_icon = QLabel(icon_text)
        lbl_icon.setStyleSheet("font-size: 32px;")
        
        lbl_title = QLabel(title)
        lbl_title.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {accent_color};")
        
        header.addWidget(lbl_icon)
        header.addWidget(lbl_title)
        header.addStretch()
        content.addLayout(header)
        
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background-color: #DEE2E6;")
        content.addWidget(line)
        
        # Message
        lbl_msg = QLabel(message_html)
        lbl_msg.setWordWrap(True)
        lbl_msg.setTextFormat(Qt.RichText)
        lbl_msg.setStyleSheet("font-size: 14px; color: #495057;")
        content.addWidget(lbl_msg)
        
        content.addStretch()
        
        # Button
        btn = QPushButton("OK")
        btn.setCursor(Qt.PointingHandCursor)
        btn.setFixedHeight(40)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {accent_color}; color: white; font-weight: bold; border-radius: 6px; border: none;
            }}
            QPushButton:hover {{ background-color: #0B5ED7; }}
        """)
        btn.clicked.connect(self.accept)
        content.addWidget(btn)

class ModernGuideDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.resize(850, 700)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        container = QFrame()
        container.setStyleSheet("background-color: #FFFFFF; border-radius: 12px; border: 1px solid #E9ECEF;")
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(25)
        shadow.setYOffset(5)
        shadow.setColor(QColor(0, 0, 0, 40))
        container.setGraphicsEffect(shadow)
        layout.addWidget(container)
        
        inner = QVBoxLayout(container)
        inner.setContentsMargins(0, 0, 0, 0)
        inner.setSpacing(0)
        
        # Header
        header = QFrame()
        header.setStyleSheet("background-color: #F8F9FA; border-top-left-radius: 12px; border-top-right-radius: 12px; border-bottom: 1px solid #DEE2E6;")
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(20, 15, 20, 15)
        
        lbl_t = QLabel("📚 User Guide & Tutorial")
        lbl_t.setStyleSheet("font-size: 18px; font-weight: bold; color: #212529;")
        
        btn_x = QPushButton("✕")
        btn_x.setCursor(Qt.PointingHandCursor)
        btn_x.setFixedSize(32, 32)
        btn_x.setStyleSheet("""
            QPushButton { 
                border: none; 
                background-color: #F8D7DA; 
                color: #721C24; 
                border-radius: 6px; 
                font-size: 16px; 
                font-weight: bold;
            } 
            QPushButton:hover { 
                background-color: #DC3545; 
                color: white; 
            }
        """)
        btn_x.clicked.connect(self.accept)
        
        h_layout.addWidget(lbl_t)
        h_layout.addStretch()
        h_layout.addWidget(btn_x)
        
        inner.addWidget(header)
        
        # Content
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setFrameShape(QFrame.NoFrame)
        self.text_edit.setStyleSheet("padding: 20px; border-bottom-left-radius: 12px; border-bottom-right-radius: 12px;")
        
        # Enhanced HTML Content
        self.text_edit.setHtml("""
            <style>
                h2 { color: #0D6EFD; margin-bottom: 15px; font-family: 'Segoe UI', sans-serif; }
                h3 { color: #0DCAF0; margin-top: 25px; margin-bottom: 10px; font-size: 16px; }
                p, li { font-size: 14px; line-height: 1.6; color: #495057; }
                .card { 
                    background-color: #F8F9FA; 
                    border: 1px solid #E9ECEF; 
                    border-radius: 8px; 
                    padding: 15px; 
                    margin-bottom: 15px; 
                }
                .highlight { color: #D63384; font-family: monospace; background: #FFF0F6; padding: 2px 5px; border-radius: 4px; }
                .step { font-weight: bold; color: #198754; }
            </style>
            
            <h2>👋 Welcome to Mail Merge Pro!</h2>
            <p>Send personalized bulk emails directly from your Gmail Drafts using Excel data.</p>
            
            <h3>1. 📊 Excel File Setup</h3>
            <div class="card">
                <p>Ensure your Excel file has a <b>Header Row</b>.</p>
                <ul>
                    <li><b>Required:</b> Use a column named <span class="highlight">Email</span>.</li>
                    <li><b>Data:</b> Add columns like <span class="highlight">Name</span>, <span class="highlight">Company</span>, etc.</li>
                    <li><b>CC/BCC:</b> Optional columns <span class="highlight">CC</span> and <span class="highlight">BCC</span> (comma-separated).</li>
                </ul>
            </div>
            
            <h3>2. 📝 Creating Templates</h3>
            <div class="card">
                <p>Create a Draft in Gmail. Use placeholders matching your Excel headers:</p>
                <p><i>"Hi <span class="highlight">{{Name}}</span>, attached is the report for <span class="highlight">{{Company}}</span>."</i></p>
                <p><b>Note:</b> You can also use placeholders in the Subject Line!</p>
            </div>
            
            <h3>3. 🚀 Sending Process</h3>
            <div class="card">
                <p><span class="step">Step 1:</span> Login with Google.</p>
                <p><span class="step">Step 2:</span> Select your Excel file.</p>
                <p><span class="step">Step 3:</span> Select your Draft.</p>
                <p><span class="step">Step 4:</span> Click <b>Start New</b> to begin!</p>
            </div>

            <h3>4. 📎 Conditional Attachments</h3>
            <div class="card">
                <p>You can choose to send attachments only to specific people.</p>
                <p>1. Uncheck the <b>"Send Attachments"</b> box (next to Start button).</p>
                <p>2. The app will look for an Excel column named <span class="highlight">Send Attachments</span>.</p>
                <ul>
                    <li><b>Yes</b> (or empty): Attachments sent.</li>
                    <li><b>No</b>: Attachments removed for that user.</li>
                </ul>
            </div>
            
            <p style="color: #6C757D; text-align: center; margin-top: 20px;"><i>Use the Stop/Resume buttons to manage large lists securely.</i></p>
        """)
        
        inner.addWidget(self.text_edit)

class ModernResultDialog(QDialog):
    def __init__(self, parent, prev_sent, prev_failed, sent, failed, total_sent, total_failed):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        # Determine strict mode (Simple vs Resume view)
        # If no previous data, show simple view
        self.is_simple_mode = (prev_sent == 0 and prev_failed == 0)
        
        height = 350 if self.is_simple_mode else 480
        self.resize(380, height) # Compact width
        
        # Main Layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Container Card
        self.container = QFrame()
        self.container.setObjectName("Container")
        self.container.setStyleSheet("""
            QFrame#Container {
                background-color: #FFFFFF;
                border-radius: 16px;
                border: 1px solid #E9ECEF;
            }
        """)
        
        # Drop Shadow
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(25)
        shadow.setXOffset(0)
        shadow.setYOffset(5)
        shadow.setColor(QColor(0, 0, 0, 40))
        self.container.setGraphicsEffect(shadow)
        
        layout.addWidget(self.container)
        
        # Content Layout
        content_layout = QVBoxLayout(self.container)
        content_layout.setSpacing(20)
        content_layout.setContentsMargins(30, 40, 30, 40)
        
        # 1. Header with Icon
        header_layout = QVBoxLayout()
        header_layout.setSpacing(15)
        
        # Professional Check Icon
        icon_container = QLabel()
        icon_container.setText("✓")
        icon_container.setAlignment(Qt.AlignCenter)
        icon_container.setFixedSize(64, 64)
        icon_container.setStyleSheet("""
            background-color: #D1E7DD; 
            color: #198754; 
            font-size: 32px; 
            font-weight: bold; 
            border-radius: 32px;
        """)
        
        # Center the icon
        icon_wrapper = QHBoxLayout()
        icon_wrapper.addStretch()
        icon_wrapper.addWidget(icon_container)
        icon_wrapper.addStretch()
        header_layout.addLayout(icon_wrapper)
        
        lbl_title = QLabel("Process Complete")
        lbl_title.setAlignment(Qt.AlignCenter)
        lbl_title.setStyleSheet("font-family: 'Segoe UI', sans-serif; font-size: 24px; font-weight: 700; color: #212529;")
        
        header_layout.addWidget(lbl_title)
        content_layout.addLayout(header_layout)
        
        # Divider
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Plain)
        line.setStyleSheet("background-color: #DEE2E6;")
        content_layout.addWidget(line)
        
        # 2. Stats Grid (Only if Resume)
        if not self.is_simple_mode:
            stats_layout = QGridLayout()
            stats_layout.setHorizontalSpacing(40)
            stats_layout.setVerticalSpacing(15)
            
            # Headers
            lbl_prev = QLabel("Before Resume")
            lbl_prev.setAlignment(Qt.AlignCenter)
            lbl_prev.setStyleSheet("font-weight: 700; color: #495057; font-size: 14px; text-transform: uppercase; letter-spacing: 0.5px;")
            
            lbl_curr = QLabel("After Resume")
            lbl_curr.setAlignment(Qt.AlignCenter)
            lbl_curr.setStyleSheet("font-weight: 700; color: #495057; font-size: 14px; text-transform: uppercase; letter-spacing: 0.5px;")
            
            stats_layout.addWidget(lbl_prev, 0, 0)
            stats_layout.addWidget(lbl_curr, 0, 1)
            
            # Helper to create value row
            def create_stat_row(label, value, is_success):
                row = QHBoxLayout()
                row.setSpacing(8)
                
                icon = QLabel("✔" if is_success else "✖")
                color = "#198754" if is_success else "#DC3545" # Bootstrap Success/Danger
                icon.setStyleSheet(f"color: {color}; font-size: 16px; font-weight: 900;")
                
                txt = QLabel(f"{label}: {value}")
                txt.setStyleSheet("color: #6C757D; font-size: 15px; font-weight: 500;")
                
                row.addStretch()
                row.addWidget(icon)
                row.addWidget(txt)
                row.addStretch()
                return row

            # Before Resume - Rows
            prev_sent_layout = create_stat_row("Sent", prev_sent, True)
            prev_fail_layout = create_stat_row("Failed", prev_failed, False)
            
            # After Resume - Rows
            curr_sent_layout = create_stat_row("Sent", sent, True)
            curr_fail_layout = create_stat_row("Failed", failed, False)

            stats_layout.addLayout(prev_sent_layout, 1, 0)
            stats_layout.addLayout(prev_fail_layout, 2, 0)
            
            stats_layout.addLayout(curr_sent_layout, 1, 1)
            stats_layout.addLayout(curr_fail_layout, 2, 1)
            
            content_layout.addLayout(stats_layout)
        else:
            # Simple Spacer if no grid
            content_layout.addSpacing(10)
        
        # 3. Total Section (Always shown, simplified if simple mode)
        total_frame = QFrame()
        total_frame.setStyleSheet("""
            QFrame {
                background-color: #F8F9FA;
                border-radius: 10px;
                border: 1px solid #E9ECEF;
            }
        """)
        total_layout = QHBoxLayout(total_frame)
        total_layout.setContentsMargins(20, 15, 20, 15)
        
        # In simple mode, we can drop the "TOTAL" text or keep it. user said "show normal sent and fails".
        # Let's keep "TOTAL" but maybe center it more?
        # Actually, if we hide the grid, this is the only stats. 
        # User requested "show normal sent and fails".
        
        lbl_total_title = QLabel("TOTAL" if not self.is_simple_mode else "SUMMARY")
        lbl_total_title.setStyleSheet("font-weight: 900; font-size: 16px; color: #212529; letter-spacing: 1px;")
        
        # Shared style for counts
        count_style = "font-weight: 500; font-size: 16px;"  if not self.is_simple_mode else "font-weight: 600; font-size: 18px;"
        
        # Icon Style
        icon_style_success = "color: #198754; font-weight: 900; font-size: 18px;"
        icon_style_fail = "color: #DC3545; font-weight: 900; font-size: 18px;"
        
        lbl_check = QLabel("✔ SENT:")
        if self.is_simple_mode: lbl_check.setText("✔ Sent:")
        lbl_check.setStyleSheet(icon_style_success)
        
        lbl_ts = QLabel(str(total_sent))
        lbl_ts.setStyleSheet(count_style + "color: #198754;")
        
        lbl_cross = QLabel("✖ FAILED:")
        if self.is_simple_mode: lbl_cross.setText("✖ Failed:")
        lbl_cross.setStyleSheet(icon_style_fail)
        
        lbl_tf = QLabel(str(total_failed))
        lbl_tf.setStyleSheet(count_style + "color: #DC3545;")
        
        if self.is_simple_mode:
            # Center everything in simple mode
            total_layout.addStretch()
            total_layout.addWidget(lbl_check)
            total_layout.addSpacing(5)
            total_layout.addWidget(lbl_ts)
            total_layout.addSpacing(30)
            total_layout.addWidget(lbl_cross)
            total_layout.addSpacing(5)
            total_layout.addWidget(lbl_tf)
            total_layout.addStretch()
        else:
            total_layout.addWidget(lbl_total_title)
            total_layout.addStretch()
            
            total_layout.addWidget(lbl_check)
            total_layout.addSpacing(5)
            total_layout.addWidget(lbl_ts)
            
            total_layout.addSpacing(25)
            
            total_layout.addWidget(lbl_cross)
            total_layout.addSpacing(5)
            total_layout.addWidget(lbl_tf)
        
        content_layout.addWidget(total_frame)
        
        content_layout.addStretch()
        
        # 4. Action Button
        btn_done = QPushButton("Done!")
        btn_done.setCursor(Qt.PointingHandCursor)
        btn_done.setFixedHeight(50)
        btn_done.setStyleSheet("""
            QPushButton {
                background-color: #0D6EFD;
                color: white;
                font-family: 'Segoe UI', sans-serif;
                font-weight: 600;
                font-size: 16px;
                border-radius: 8px;
                border: none;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #0B5ED7;
            }
            QPushButton:pressed {
                background-color: #0A58CA;
            }
        """)
        btn_done.clicked.connect(self.accept)
        content_layout.addWidget(btn_done)


# --- LOADING OVERLAY & WORKER ---

class LoadingOverlay(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False) # Catch clicks
        self.setAttribute(Qt.WA_NoSystemBackground)
        self.setHidden(True)
        
        # Loader Animation Variables
        self.angle = 0
        self.span = 0
        self.growing = True
        
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.rotate)
        
        # Styles
        self.bg_color = QColor(255, 255, 255, 240) # Slightly more opaque white
        self.spinner_color = QColor("#0D6EFD") # Bootstrap Primary Blue
        self.loading_text = "Processing..."

    def rotate(self):
        # Rotate the whole spinner
        self.angle = (self.angle + 10) % 360
        
        # "Snake" effect: span grows and shrinks
        if self.growing:
            self.span += 4
            if self.span >= 260: # Max arc length
                self.growing = False
        else:
            self.span -= 4
            if self.span <= 20: # Min arc length
                self.growing = True
                
        self.update()

    def show_loading(self, text="Loading..."):
        self.loading_text = text
        self.setHidden(False)
        self.raise_()
        self.timer.start(16) # ~60fps for smoothness
        if self.parent():
            self.resize(self.parent().size())

    def hide_loading(self):
        self.setHidden(True)
        self.timer.stop()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # 1. Overlay Background
        painter.setBrush(self.bg_color)
        painter.setPen(Qt.NoPen)
        painter.drawRect(self.rect())
        
        # 2. Spinner (Google-style Snake)
        width = 60
        height = 60
        # Center coordinates
        x = (self.width() - width) // 2
        y = (self.height() - width) // 2 - 20 # Slightly above center
        
        rect = QRect(x, y, width, height)
        pen = QPen(self.spinner_color)
        pen.setWidth(6)
        pen.setCapStyle(Qt.RoundCap)
        painter.setPen(pen)
        
        # Draw the "Snake" Arc
        # 'angle' rotates the frame, 'span' changes the length
        # Start angle must also move to keep visual momentum when shrinking
        # Simplification: Draw arc starting at 'self.angle' with length 'self.span'
        # For true Google material effect, the start angle moves faster when shrinking.
        # But a simple grow/shrink with constant rotation is a good approximation.
        
        painter.drawArc(rect, -self.angle * 16, -self.span * 16)
        
        # 3. Loading Text
        painter.setPen(QColor("#333333"))
        font = QFont("Segoe UI", 12)
        font.setWeight(QFont.Bold)
        painter.setFont(font)
        
        text_rect = QRect(0, y + height + 15, self.width(), 40)
        painter.drawText(text_rect, Qt.AlignCenter, self.loading_text)
        
    def resizeEvent(self, event):
        # Always cover parent
        if self.parent():
            self.resize(self.parent().size())
        super().resizeEvent(event)


class StartupWorker(QThread):
    # Signals
    status_signal = pyqtSignal(str)
    log_signal = pyqtSignal(str, str) # msg, color
    auth_success = pyqtSignal(object, object, dict) # creds, service, user_info
    drafts_loaded = pyqtSignal(list, list) # draft_list, raw_drafts(optional - internal usage)
    error_signal = pyqtSignal(str)
    
    def __init__(self, force_auth=False):
        super().__init__()
        self.force_auth = force_auth
        
    def run(self):
        try:
            # 1. AUTHENTICATION
            self.status_signal.emit("Authenticating with Google...")
            
            creds = None
            creds_file = resource_path('token.json')
            cred_json = resource_path('credentials.json')
            
            if os.path.exists(creds_file):
                creds = Credentials.from_authorized_user_file(creds_file, SCOPES)
            
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    self.log_signal.emit("Refreshing expired token...", "#17A2B8")
                    creds.refresh(Request())
                else:
                    if self.force_auth:
                        self.log_signal.emit("Initiating new login flow...", "#0D6EFD")
                        flow = InstalledAppFlow.from_client_secrets_file(cred_json, SCOPES)
                        creds = flow.run_local_server(port=0, prompt='select_account')
                        with open(creds_file, 'w') as token:
                            token.write(creds.to_json())
                    else:
                        # Fail silently if not forced; return early
                        self.log_signal.emit("Authentication required - please login.", "#DC3545")
                        return

            # Build Service
            service = build('gmail', 'v1', credentials=creds)
            
            # 2. FETCH PROFILE
            self.status_signal.emit("Fetching user profile...")
            profile = service.users().getProfile(userId='me').execute()
            user_info_oauth = build('oauth2', 'v2', credentials=creds).userinfo().get().execute()
            
            user_data = {
                'email': profile.get('emailAddress'),
                'name': user_info_oauth.get('name'),
                'picture': user_info_oauth.get('picture'),
                'avatar_bytes': None
            }
            
            # Optimization: Fetch Avatar Bytes in background
            if user_data['picture']:
                try:
                    user_data['avatar_bytes'] = requests.get(user_data['picture']).content
                except:
                    pass
            
            # Emit Auth Success
            self.auth_success.emit(creds, service, user_data)
            
            # 3. LOAD DRAFTS
            self.status_signal.emit("Loading Gmail drafts...")
            
            drafts = service.users().drafts().list(userId='me').execute().get('drafts', [])
            
            formatted_drafts = []
            for d in drafts:
                # We need details for the UI. Fetching details is network heavy.
                detail = service.users().drafts().get(userId='me', id=d['id']).execute()
                headers = detail['message']['payload']['headers']
                subject = next((h['value'] for h in headers if h['name'] == 'Subject'), '(No Subject)')
                formatted_drafts.append(f"{subject} [ID: {d['id']}]")
            
            self.drafts_loaded.emit(formatted_drafts, drafts)
            
            self.status_signal.emit("Ready!")
            
        except Exception as e:
            self.error_signal.emit(str(e))
            self.log_signal.emit(f"Startup Error: {e}", "#DC3545")


class SkeletonItem(QWidget):
    def __init__(self, width=None, height=None, shape="box", parent=None):
        super().__init__(parent)
        if width: self.setFixedWidth(width)
        if height: self.setFixedHeight(height)
        self.shape = shape
        
        # Animation: Shimmer from Left to Right
        self.slide = 0.0 # Progress from 0.0 to 1.0
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.animate)
        self.timer.start(20) # Faster timer for smooth movement

    def animate(self):
        # Move the shimmer wave
        self.slide += 0.02
        if self.slide > 1.5: # Go a bit past the end for a pause effect
            self.slide = -0.5 # Start from before the beginning
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        rect = self.rect()
        w = rect.width()
        h = rect.height()
        
        # Create Linear Gradient for Shimmer
        # We simulate a light source moving across.
        # The gradient needs to be wider than the item to see the "beam" move.
        
        # Calculate current position of the "beam"
        pos = self.slide * w
        
        gradient = QLinearGradient(0, 0, w, 0)
        gradient.setSpread(QGradient.PadSpread)
        
        # Base Color: Grey (#DEE2E6)
        # Highlight: White (#FFFFFF)
        grey = QColor("#DEE2E6")
        white = QColor("#FFFFFF")
        
        # We need the gradient coordinates to move, or the stops to move.
        # It's easier to define the gradient relative to the widget and simple stops,
        # but for a moving "beam", defining coordinates relative to 'pos' is better.
        
        # Let's try a gradient that spans from (pos - beam_width) to (pos + beam_width)
        beam_width = w * 0.6 # The beam is 60% of element width
        
        shimmer_gradient = QLinearGradient(pos - beam_width, 0, pos + beam_width, 0)
        shimmer_gradient.setColorAt(0.0, grey)
        shimmer_gradient.setColorAt(0.5, white) # Center of beam is white
        shimmer_gradient.setColorAt(1.0, grey)
        
        painter.setBrush(QBrush(shimmer_gradient))
        painter.setPen(Qt.NoPen)
        
        if self.shape == "circle":
            painter.drawEllipse(rect)
        else:
            painter.drawRoundedRect(rect, 8, 8)

class SkeletonWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # 1. Header (Logo)
        header = QHBoxLayout()
        header.addStretch()
        header.addWidget(SkeletonItem(200, 80)) # Logo placeholder
        header.addStretch()
        layout.addLayout(header)
        
        # 2. Profile Card
        card1 = QFrame()
        card1.setStyleSheet("background-color: white; border-radius: 12px; border: 1px solid #E9ECEF;")
        c1_layout = QHBoxLayout(card1)
        
        # Avatar
        c1_layout.addWidget(SkeletonItem(50, 50, "circle"))
        # Name
        c1_layout.addWidget(SkeletonItem(200, 20))
        c1_layout.addStretch()
        # Buttons
        c1_layout.addWidget(SkeletonItem(120, 40))
        c1_layout.addWidget(SkeletonItem(150, 40))
        
        layout.addWidget(card1)
        
        # 3. Body (Draft Selection)
        card2 = QFrame()
        card2.setStyleSheet("background-color: white; border-radius: 12px; border: 1px solid #E9ECEF;")
        c2_layout = QVBoxLayout(card2)
        c2_layout.addWidget(SkeletonItem(None, 200)) # Big box
        layout.addWidget(card2)
        
        # 4. Footer (Action Buttons)
        footer = QHBoxLayout()
        footer.addStretch()
        footer.addWidget(SkeletonItem(150, 50))
        footer.addWidget(SkeletonItem(150, 50))
        footer.addStretch()
        layout.addLayout(footer)
        
        # 5. Log
        layout.addWidget(SkeletonItem(None, 100))
        layout.addStretch()


class ModernProgressBar(QProgressBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(12) # Reduced by 60% from 30px
        self.setTextVisible(False)
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        rect = self.rect()
        width = rect.width()
        height = rect.height()
        
        # 1. Background (Unfilled)
        painter.setBrush(QColor("#E9ECEF"))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(rect, 6, 6) # Adjusted radius
        
        # 2. Filled Portion (Left to Right)
        val = self.value()
        mx = self.maximum()
        if mx == 0: mx = 1
        ratio = val / mx
        
        fill_width = int(width * ratio)
        if fill_width > 0:
            fill_rect = QRect(0, 0, fill_width, height)
            
            # Solid Color #0A85B5
            painter.setBrush(QColor("#0A85B5"))
            
            painter.save()
            painter.setClipRect(fill_rect) 
            painter.drawRoundedRect(rect, 6, 6)
            painter.restore()
            
        # 3. Text (Centered, White)
        # Note: 12px is very thin for text. Using small font.
        text = f"{int(ratio * 100)}%"
        painter.setPen(QColor("white"))
        font = QFont("Segoe UI", 8) 
        font.setWeight(QFont.Bold)
        painter.setFont(font)
        
        # Text Rendering
        painter.drawText(rect, Qt.AlignCenter, text)


# --- MAIN APPLICATION WINDOW ---
class MailMergeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Mail Merge Pro 14.0")
        self.resize(1100, 850)
        
        # Set Window Icon
        try:
            icon_path = resource_path("app_icon.ico") 
            if os.path.exists(icon_path):
                self.setWindowIcon(QIcon(icon_path))
        except: pass
        
        # Variables
        self.creds = None
        self.service = None
        self.excel_path = ""
        self.drafts = []
        self.worker = None
        self.preview_header_map = {} # Map header name -> col index
        
        # Track Cumulative Stats
        self.total_sent = 0
        self.total_failed = 0
        
        # Load Creds
        self.valid_header_indices = [] # Indices of non-empty headers
        self.live_preview_buffer = [] # List of tuples: (row_idx, row_values, status)

        self.init_ui()
        self.init_menu()
        
        # Delayed auto-auth to let UI show up
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(100, self.auto_authenticate)

    def init_ui(self):
        # --- GLOBAL STYLESHEET (Standard Clean Theme) ---
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F0F2F5; /* Light Grey background */
            }
            QLabel {
                color: #1C1E21;
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;
            }
            QFrame#Card {
                background-color: #FFFFFF;
                border-radius: 12px;
                border: 1px solid #DCE0E5;
            }
            QListWidget, QTextEdit {
                background-color: #FFFFFF;
                border-radius: 8px;
                border: 1px solid #DCE0E5;
                color: #333333;
                padding: 10px;
                font-family: 'Consolas', monospace;
            }
        """)

        # Central Widget acts as a Switcher (Stack)
        central = QWidget()
        self.setCentralWidget(central)
        self.stack = QStackedLayout(central)
        
        # View 0: Skeleton Loader
        self.skeleton = SkeletonWidget()
        self.stack.addWidget(self.skeleton)
        
        # View 1: Main Content
        self.main_content_widget = QWidget()
        main_layout = QVBoxLayout(self.main_content_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)
        self.stack.addWidget(self.main_content_widget) # Index 1

        # --- HEADER AREA (Logo Only) ---
        header_layout = QHBoxLayout()
        header_layout.addStretch() # Center the logo? Or keep left? User didn't specify. 
        # Previous state had logo left. Let's keep it simple.
        # Actually, let's just use the logo code directly if we don't need a complex header layout anymore.
        # But keeping the layout is safer for future changes.
        
        try:
            logo_path = resource_path("logo.png")
            if os.path.exists(logo_path):
                lbl_logo = QLabel()
                pixmap = QPixmap(logo_path).scaledToHeight(80, Qt.SmoothTransformation) # Slightly larger logo
                lbl_logo.setPixmap(pixmap)
                lbl_logo.setAlignment(Qt.AlignCenter)
                header_layout.addWidget(lbl_logo)
        except: pass
        
        header_layout.addStretch()
        main_layout.addLayout(header_layout)

        # --- CARD 1: PROFILE & ACTIONS ---
        card1 = QFrame()
        card1.setObjectName("Card")
        # Main layout for the card is Vertical to hold the "Top Row" and "Preview Table"
        card1_main_layout = QVBoxLayout(card1) 
        card1_main_layout.setSpacing(15)

        # --- Top Row: Profile (Left) | Space | Buttons (Right) ---
        top_row = QHBoxLayout()

        # 1. Profile Section (Left)
        profile_box = QHBoxLayout()
        profile_box.setSpacing(10)
        
        self.lbl_avatar = QLabel()
        self.lbl_avatar.setFixedSize(50, 50)
        self.lbl_avatar.setStyleSheet("border-radius: 25px; background-color: #E9ECEF; border: 1px solid #DEE2E6;")
        profile_box.addWidget(self.lbl_avatar)
        
        self.lbl_user = QLabel("Not logged in")
        self.lbl_user.setStyleSheet("color: #333; font-weight: bold; font-size: 16px;")
        profile_box.addWidget(self.lbl_user)
        
        top_row.addLayout(profile_box)
        
        # 2. Spacer (Space Between)
        top_row.addStretch()

        # 3. Buttons Section (Right)
        btns_box = QHBoxLayout()
        
        self.btn_auth = QPushButton("🔐 Login / Re-Auth")
        self.style_standard_button(self.btn_auth, (23, 162, 184)) # Cyan
        self.btn_auth.clicked.connect(self.manual_authenticate)
        btns_box.addWidget(self.btn_auth)
        
        btns_box.addSpacing(10)

        self.btn_excel = QPushButton("📁 Select Excel File")
        self.style_standard_button(self.btn_excel, (23, 162, 184)) # Cyan
        self.btn_excel.clicked.connect(self.choose_excel)
        btns_box.addWidget(self.btn_excel)
        
        # Sync Button
        self.btn_sync = QPushButton(" Refresh Excel")
        self.btn_sync.setIcon(QIcon(resource_path("refresh.png")))
        self.btn_sync.setIconSize(QSize(20, 20))
        self.style_standard_button(self.btn_sync, (23, 162, 184)) # #17A2B8
        self.btn_sync.setToolTip("Reload Excel File")
        self.btn_sync.clicked.connect(self.reload_excel)
        btns_box.addWidget(self.btn_sync)
        
        top_row.addLayout(btns_box)
        
        card1_main_layout.addLayout(top_row)

        # 3. Preview Table (Hidden by default)
        self.table_preview = QTableWidget()
        self.table_preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_preview.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_preview.setAlternatingRowColors(True)
        self.table_preview.verticalHeader().setVisible(False)
        # Allow horizontal scrolling
        self.table_preview.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_preview.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table_preview.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                gridline-color: #E9ECEF;
                border: 1px solid #DEE2E6;
                border-radius: 6px;
            }
            QHeaderView::section {
                background-color: #F8F9FA;
                padding: 4px;
                border: 1px solid #DEE2E6;
                font-weight: bold;
                color: #495057;
            }
        """)
        self.table_preview.setVisible(False) # Hide initially
        
        # Hide space initially
        self.table_preview.setFixedHeight(0) 
        
        self.table_preview.setCursor(Qt.PointingHandCursor)
        
        # Force Scrollbars as requested ("X & Y axis scrollbar should be there")
        self.table_preview.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.table_preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        
        card1_main_layout.addWidget(self.table_preview)

        main_layout.addWidget(card1)

        # --- CARD 2: GMAIL DRAFT SELECTION ---
        card2 = QFrame()
        card2.setObjectName("Card")
        card2_layout = QVBoxLayout(card2)
        
        lbl_draft = QLabel("✉️ Select Gmail Draft")
        lbl_draft.setStyleSheet("font-weight: bold; font-size: 16px; color: #0DCAF0;") # Cyan Title
        
        # Drafts List with Refresh
        drafts_row = QHBoxLayout()
        drafts_row.addWidget(lbl_draft)
        drafts_row.addStretch()
        
        btn_reload_drafts = QPushButton()
        btn_reload_drafts.setIcon(QIcon(resource_path("refresh.png")))
        self.style_icon_button(btn_reload_drafts)
        btn_reload_drafts.setToolTip("Reload Drafts")
        btn_reload_drafts.clicked.connect(self.load_drafts)
        drafts_row.addWidget(btn_reload_drafts)
        
        card2_layout.addLayout(drafts_row)
        
        self.list_drafts = QListWidget()
        self.list_drafts.setFixedHeight(120)
        self.list_drafts.setSelectionMode(QAbstractItemView.SingleSelection)
        card2_layout.addWidget(self.list_drafts)
        
        main_layout.addWidget(card2)

        # --- CARD 3: ACTION BUTTONS (Start, Resume, Stop) ---
        card3 = QFrame()
        card3.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 1px solid #DCE0E5; }")
        card3_layout = QHBoxLayout(card3)
        card3_layout.setContentsMargins(20, 20, 20, 20)
        card3_layout.setSpacing(20)
        
        card3_layout.addStretch()
        
        # Attachment Toggle
        self.chk_send_attachments = QCheckBox("Send Attachments")
        self.chk_send_attachments.setChecked(True) # Default: Send to all
        self.chk_send_attachments.setCursor(Qt.PointingHandCursor)
        self.chk_send_attachments.setStyleSheet("font-weight: bold; color: #495057; font-size: 14px;")
        self.chk_send_attachments.clicked.connect(self.on_toggle_attachments)
        card3_layout.addWidget(self.chk_send_attachments)
        
        card3_layout.addSpacing(15)
        
        # Start New Button
        self.btn_process = QPushButton("🚀 Start New")
        self.style_standard_button(self.btn_process, (40, 167, 69)) # Green
        self.btn_process.clicked.connect(self.start_mail_merge)
        card3_layout.addWidget(self.btn_process)
        
        # Resume Button
        self.btn_resume = QPushButton("▶ Resume")
        self.style_standard_button(self.btn_resume, (255, 193, 7)) # Yellow/Amber
        self.btn_resume.clicked.connect(self.resume_mail_merge)
        self.btn_resume.setEnabled(False) # Initially disabled
        card3_layout.addWidget(self.btn_resume)
        
        # Stop Button
        self.btn_stop = QPushButton("⏹ STOP")
        self.style_standard_button(self.btn_stop, (220, 53, 69)) # Red
        self.btn_stop.setStyleSheet(self.btn_stop.styleSheet().replace("background-color: #E0E0E0;", "background-color: white; border: 1px solid #DC3545; color: #DC3545;")) # Outline style for stop
        self.btn_stop.clicked.connect(self.stop_mail_merge)
        self.btn_stop.setEnabled(False)
        card3_layout.addWidget(self.btn_stop)
        
        card3_layout.addStretch()
        
        main_layout.addWidget(card3)

        # --- PROGRESS BAR ---
        self.progress_bar = ModernProgressBar()
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)
        
        # --- LOG CONSOLE ---
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFixedHeight(80) # Reduced to ~3 lines
        self.txt_log.setPlaceholderText("System logs will appear here...")
        main_layout.addWidget(self.txt_log)
        
        # Footer
        footer_layout = QHBoxLayout()
        footer_layout.addStretch()
        lbl_copy = QLabel("@Copyrights2025 by Balvant Sharma. All rights reserved.")
        lbl_copy.setStyleSheet("color: #ADB5BD; font-size: 12px;")
        footer_layout.addWidget(lbl_copy)
        footer_layout.addStretch()
        
        main_layout.addStretch() # Push everything up
        main_layout.addLayout(footer_layout)
        
        self.overlay = LoadingOverlay(self) # Initialize overlay

    def resizeEvent(self, event):
        if hasattr(self, 'overlay'):
            self.overlay.resize(self.size()) # Ensure overlay covers parent
        super().resizeEvent(event)

    def init_menu(self):
        menubar = self.menuBar()
        
        # File Menu
        file_menu = menubar.addMenu('File')
        exit_action = QAction('Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Help Menu
        help_menu = menubar.addMenu('Help')
        
        guide_action = QAction('How to Use', self)
        guide_action.triggered.connect(self.show_guide)
        help_menu.addAction(guide_action)
        
        contact_action = QAction('Contact to Developer', self)
        contact_action.triggered.connect(self.show_contact_info)
        help_menu.addAction(contact_action)

    def show_contact_info(self):
        msg = "<b>Name:</b> Balvant Sharma<br><b>Email:</b> balavantsharma91@gmail.com"
        dlg = ModernInfoDialog(self, "Developer Contact", msg, "👨‍💻", "#17A2B8")
        dlg.exec_()

    def show_guide(self):
        dlg = ModernGuideDialog(self)
        dlg.exec_()

    def style_standard_button(self, btn, rgb):
        r, g, b = rgb
        # Calculate darker hover shade
        darker_r = max(0, r - 30)
        darker_g = max(0, g - 30)
        darker_b = max(0, b - 30)
        
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: rgb({r}, {g}, {b});
                color: white;
                border: none;
                border-radius: 8px; 
                padding: 10px 20px;
                font-weight: bold;
                font-size: 13px;
                min-width: 120px;
            }}
            QPushButton:hover {{
                background-color: rgb({darker_r}, {darker_g}, {darker_b}); /* Solid darker hover */
            }}
            QPushButton:pressed {{
                background-color: rgba({r}, {g}, {b}, 150);
            }}
            QPushButton:disabled {{
                background-color: #E0E0E0;
                color: #A0A0A0;
            }}
        """)

    def style_icon_button(self, btn):
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #E9ECEF;
                border-radius: 5px;
            }
            QPushButton:pressed {
                background-color: #DEE2E6;
            }
        """)

    def apply_progress_style(self, color_hex):
        # Deprecated: ModernProgressBar handles its own painting.
        # Keeping empty method to prevent crashes from existing calls.
        pass


    # --- LOGIC METHODS ---
    def log(self, msg, color="black"):
        html = f"<span style='color:{color}'>{msg}</span>"
        self.txt_log.moveCursor(QTextCursor.Start)
        self.txt_log.insertHtml(html + "<br>")
        self.txt_log.moveCursor(QTextCursor.Start) 

    def update_progress(self, val):
        self.progress_bar.setValue(val)

    def manual_authenticate(self):
        if "Sign Out" in self.btn_auth.text():
            self.logout()
        else:
            creds_file = 'token.json'
            if os.path.exists(creds_file):
                os.remove(creds_file)
            self.auto_authenticate(force=True)

    def logout(self):
        try:
            creds_file = resource_path('token.json')
            if os.path.exists(creds_file):
                os.remove(creds_file)
            
            self.creds = None
            self.service = None
            self.user_email = None
            self.display_name = None
            
            # Reset UI
            self.lbl_user.setText("Not logged in")
            self.lbl_user.setStyleSheet("color: #333; font-weight: bold; font-size: 16px;")
            self.lbl_avatar.clear()
            self.lbl_avatar.setStyleSheet("border-radius: 25px; background-color: #E9ECEF; border: 1px solid #DEE2E6;")
            
            self.btn_auth.setText("🔐 Sign In")
            self.style_standard_button(self.btn_auth, (23, 162, 184)) # Cyan
            
            self.log("🔒 Signed out successfully.", "#6C757D")
        except Exception as e:
            self.log(f"Logout Error: {e}", "#DC3545")

    def auto_authenticate(self, force=False):
        # Show Skeleton Loader (Index 0)
        if hasattr(self, 'stack'):
            self.stack.setCurrentIndex(0)
        
        # Disable buttons during load
        self.btn_auth.setEnabled(False)
        self.btn_excel.setEnabled(False)
        
        self.startup_worker = StartupWorker(force_auth=force)
        # self.startup_worker.status_signal.connect(lambda s: self.overlay.show_loading(s)) # Disabled for Skeleton View
        self.startup_worker.log_signal.connect(self.log)
        self.startup_worker.auth_success.connect(self.on_startup_auth_success)
        self.startup_worker.drafts_loaded.connect(self.on_startup_drafts_loaded)
        self.startup_worker.error_signal.connect(self.on_startup_error)
        self.startup_worker.finished.connect(self.on_startup_finished)
        self.startup_worker.start()

    def on_startup_auth_success(self, creds, service, user_data):
        self.creds = creds
        self.service = service
        self.display_name = user_data.get('name')
        self.user_email = user_data.get('email')
        
        # Update Profile UI
        self.lbl_user.setText(f"{self.display_name}")
        self.lbl_user.setStyleSheet("color: #0D6EFD; font-weight: bold; font-size: 20px; margin-bottom: 10px;")
        
        self.log(f"✅ Authenticated: {self.display_name}", "#0D6EFD")
        
        # Avatar
        if user_data.get('avatar_bytes'):
            try:
                pixmap = QPixmap()
                pixmap.loadFromData(user_data['avatar_bytes'])
                
                size = 50
                rounded = QPixmap(size, size)
                rounded.fill(Qt.transparent)
                painter = QPainter(rounded)
                painter.setRenderHint(QPainter.Antialiasing)
                painter.setBrush(QBrush(pixmap.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)))
                painter.setPen(QPen(QColor("#DEE2E6"), 1))
                painter.drawEllipse(0, 0, size-1, size-1)
                painter.end()
                
                self.lbl_avatar.setPixmap(rounded)
            except: pass
            
        # UI State
        self.btn_auth.setText("🔓 Sign Out")
        self.style_standard_button(self.btn_auth, (220, 53, 69)) # Red

    def on_startup_drafts_loaded(self, formatted_list, raw_drafts):
        self.list_drafts.clear()
        
        # Convert raw list to Dict {FormattedString: ID} for lookup
        self.drafts = {}
        if isinstance(raw_drafts, list):
            for i, fmt in enumerate(formatted_list):
                # Ensure index matches (it should as per StartupWorker logic)
                if i < len(raw_drafts):
                   self.drafts[fmt] = raw_drafts[i]['id']
        else:
            # Fallback if somehow already dict (unlikely with current worker)
            self.drafts = raw_drafts
            
        for item in formatted_list:
            self.list_drafts.addItem(item)
        self.log("🔄 Drafts loaded.", "#007BFF")

    def on_startup_error(self, msg):
        self.log(f"Initialization Error: {msg}", "#DC3545")
        
    def on_startup_finished(self):
        # Switch to Main Content (Index 1)
        if hasattr(self, 'stack'):
            self.stack.setCurrentIndex(1)
            
        self.btn_auth.setEnabled(True)
        self.btn_excel.setEnabled(True)
        self.startup_worker = None

    def get_user_info(self):
        try:
            profile = self.service.users().getProfile(userId='me').execute()
            self.user_email = profile.get('emailAddress')
            user_info = build('oauth2', 'v2', credentials=self.creds).userinfo().get().execute()
            self.display_name = user_info.get('name')
            
            self.lbl_user.setText(f"{self.display_name}")
            self.lbl_user.setStyleSheet("color: #007BFF; font-weight: bold; font-size: 20px; margin-bottom: 10px;")
            self.log(f"✅ Authenticated: {self.display_name}", "#007BFF")
            
            # Load Profile Picture
            photo_url = user_info.get('picture')
            if photo_url:
                try:
                    data = requests.get(photo_url).content
                    pixmap = QPixmap()
                    pixmap.loadFromData(data)
                    
                    # Circular Mask (Header Size)
                    size = 50
                    rounded = QPixmap(size, size)
                    rounded.fill(Qt.transparent)
                    painter = QPainter(rounded)
                    painter.setRenderHint(QPainter.Antialiasing)
                    painter.setBrush(QBrush(pixmap.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)))
                    painter.setPen(QPen(QColor("#DEE2E6"), 1)) # Thinner border
                    painter.drawEllipse(0, 0, size-1, size-1)
                    painter.end()
                    
                    self.lbl_avatar.setPixmap(rounded)
                except Exception as e:
                    self.log(f"⚠️ Failed to load avatar: {e}", "#FFC107")

            # Change Login button to Red (Sign Out) on successful auth
            self.btn_auth.setText("🔓 Sign Out")
            self.style_standard_button(self.btn_auth, (220, 53, 69)) # Red
        except: pass

    def load_drafts(self):
        if not self.service: return
        try:
            self.list_drafts.clear()
            self.drafts = self.service.users().drafts().list(userId='me').execute().get('drafts', [])
            for d in self.drafts:
                detail = self.service.users().drafts().get(userId='me', id=d['id']).execute()
                headers = detail['message']['payload']['headers']
                subject = next((h['value'] for h in headers if h['name'] == 'Subject'), '(No Subject)')
                self.list_drafts.addItem(f"{subject} [ID: {d['id']}]")
            self.log("🔄 Drafts refreshed.", "#007BFF")
        except Exception as e:
            self.log(f"Error loading drafts: {e}", "#DC3545")

    def choose_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel", "", "Excel Files (*.xlsx)")
        if path:
            self.excel_path = path
            filename = os.path.basename(path)
            self.btn_excel.setText(f"Selected: {filename}")
            
            # Initial Check of Headers for Attachments
            # We want to Auto-Uncheck if "Send Attachments" exists
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                # Get headers
                row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                # Check for column
                found = False
                name_found = ""
                if row1:
                    headers = [str(c).strip().lower() for c in row1 if c]
                    for name in ['send attachments', 'send attachment', 'attachment', 'attachments']:
                        if name in headers:
                            found = True
                            name_found = name
                            break
                
                # Logic: If found -> Conditional Mode -> Uncheck Box
                if found:
                    self.chk_send_attachments.setChecked(False)
                    self.log(f"📋 Found '{name_found}' column - Auto-enabled Conditional Mode", "#17A2B8")
                else:
                    self.chk_send_attachments.setChecked(True)
                    self.log(f"📂 File selected: {os.path.basename(path)}", "#007BFF")
                    self.log(f"ℹ️ Hints: Columns found: {headers}", "#6C757D")
                
                wb.close()
            except:
                self.log(f"📂 File selected (Read Error): {os.path.basename(path)}", "#FFC107")
            
            self.load_excel_data()

    def reload_excel(self):
        if hasattr(self, 'excel_path') and self.excel_path and os.path.exists(self.excel_path):
            self.load_excel_data()
            self.log("🔄 Excel reloaded.", "#17A2B8")
        else:
            self.log("⚠️ No Excel file selected to sync.", "#FFC107")

    def load_excel_data(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            
            # 1. Headers
            raw_headers = [cell.value for cell in ws[1]]
            
            # Filter empty headers
            self.valid_header_indices = [i for i, h in enumerate(raw_headers) if h and str(h).strip()]
            filtered_headers = [raw_headers[i] for i in self.valid_header_indices]
            
            # Map for Status updates
            self.preview_header_map = {h: i for i, h in enumerate(filtered_headers)}
            
            # Setup Table
            self.table_preview.setColumnCount(len(filtered_headers))
            self.table_preview.setHorizontalHeaderLabels([str(h) for h in filtered_headers])
            # Force Scrollbars (Consistent with init_ui)
            self.table_preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
            self.table_preview.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
            self.table_preview.setRowCount(0) # Clear visual data
            self.table_preview.setVisible(False) # Hide table
            self.table_preview.setFixedHeight(0) # Collapse space
            
            # Clear Buffer (Don't show data yet)
            self.live_preview_buffer = []
            
            self.log(f"📊 Excel loaded. Preview will start when sending begins.", "#17A2B8")
            
        except Exception as e:
            self.log(f"⚠️ Excel load failed: {e}", "#FFC107")
            self.log(f"⚠️ Excel load failed: {e}", "#FFC107")
            self.table_preview.setVisible(False)

    def start_mail_merge(self):
        # 1. Validate Draft Selection
        if not self.list_drafts.currentItem():
            ModernInfoDialog(self, "No Draft Selected", "Please select a Gmail draft to send.", "⚠️", "#FFC107").exec_()
            return
            
        # 2. Validate Excel Selection
        if not hasattr(self, 'excel_path') or not self.excel_path:
            ModernInfoDialog(self, "No Excel File", "Please select an Excel file with recipients.", "⚠️", "#FFC107").exec_()
            return

        # 3. Validate Conditional Attachments (NEW)
        if not self.chk_send_attachments.isChecked():
            # Check for empty cells if we are in conditional mode
            if not hasattr(self, 'attachment_empty_rule') or self.attachment_empty_rule == "yes": # Re-check even if default yes
                 # We want to catch users who maybe didn't trigger the toggle logic or modified file
                 empty_rows = self.scan_excel_for_empty_attachments()
                 if empty_rows:
                     # Show the dialog phase 2 directly
                     dlg = ModernAttachmentDialog(self)
                     dlg.show_phase_2(empty_rows)
                     if dlg.exec_() == QDialog.Accepted:
                         self.attachment_empty_rule = dlg.result_rule
                     else:
                         return # Cancelled

        # Start Loader for Preview
        self.overlay.show_loading("Loading Data for Preview...")
        self.btn_process.setEnabled(False)
        
        # Get Draft ID
        # self.drafts is dict { "Subject": "ID" }
        draft_text = self.list_drafts.currentItem().text()
        draft_id = self.drafts.get(draft_text)
        
        self.loader = DataLoadingWorker(self.service, draft_id, self.excel_path)
        self.loader.data_loaded.connect(self.show_email_preview)
        self.loader.error_signal.connect(self.on_loader_error)
        self.loader.start()

    def on_loader_error(self, err_msg):
        self.overlay.hide_loading()
        self.btn_process.setEnabled(True)
        ModernInfoDialog(self, "Error Loading Data", f"Failed to prepare preview:\n{err_msg}", "❌", "#DC3545").exec_()

    def show_email_preview(self, draft_data, wb, all_headers, visible_headers, rows):
        self.overlay.hide_loading()
        self.btn_process.setEnabled(True)
        
        # Determine CC/BCC modes initially (Ask User)
        cc_mode, global_cc = self.ask_cc_bcc("CC")
        bcc_mode, global_bcc = self.ask_cc_bcc("BCC")
         
        # Open Preview Dialog
        dlg = AdvancedPreviewDialog(self, draft_data, all_headers, visible_headers, rows, 
                                    cc_mode, global_cc, bcc_mode, global_bcc,
                                    attachment_mode=self.chk_send_attachments.isChecked())
        
        dlg.start_sending.connect(lambda: self.on_preview_confirmed(draft_data['id'], cc_mode, global_cc, bcc_mode, global_bcc))
        dlg.exec_()

    def on_preview_confirmed(self, draft_id, cc_mode, global_cc, bcc_mode, global_bcc):
        # This is called when user clicks "Start Sending" in Preview Dialog
        
        # Prepare arguments for real start
        self.pending_send_args = {
             'service': self.service,
             'excel_path': self.excel_path,
             'draft_id': draft_id,
             'start_row': 2, # Always start from 2 (Fresh)
             'cc_mode': cc_mode,
             'global_cc': global_cc,
             'bcc_mode': bcc_mode,
             'global_bcc': global_bcc,
             'display_name': self.display_name,
             'user_email': self.user_email,
             'total_rows': None 
        }
        
        # Add Attachment Rule logic if needed
        if hasattr(self, 'attachment_empty_rule'):
             self.pending_send_args['attachment_empty_rule'] = self.attachment_empty_rule
        else:
             self.pending_send_args['attachment_empty_rule'] = 'yes'
             
        self.real_start_sending()

    def real_start_sending(self):
        # Actual Worker Start
        args = self.pending_send_args
        
        self.log("🚀 Starting Mail Merge...", "#0D6EFD")
        self.btn_process.setEnabled(False)
        self.btn_resume.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_stop.setStyleSheet(self.btn_stop.styleSheet().replace("border: 1px solid #DC3545; color: #DC3545;", "background-color: #DC3545; color: white;"))
        
        # Reset Stats
        self.total_sent = 0
        self.total_failed = 0
        self.progress_bar.setValue(0)
        self.apply_progress_style("#0D6EFD") # Blue
        
        # Clear/Init Preview Table
        self.live_preview_buffer = []
        self.table_preview.setRowCount(0)
        self.table_preview.setVisible(False) 
        
        self.worker = EmailWorker(
            args['service'], args['excel_path'], args['draft_id'], args['start_row'],
            args['cc_mode'], args['global_cc'], args['bcc_mode'], args['global_bcc'],
            args['display_name'], args['user_email'], args.get('total_rows'), args.get('is_resume', False),
            attachment_mode=self.chk_send_attachments.isChecked(),
            attachment_empty_rule=args.get('attachment_empty_rule', 'yes')
        )
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.live_preview_signal.connect(self.handle_live_preview_update)
        self.worker.stopped_signal.connect(self.on_stopped_stats)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error_signal.connect(self.on_error)
        self.worker.start()

    def resume_mail_merge(self):
        self.start_sending(resume=True)

    def stop_mail_merge(self):
        if hasattr(self, 'stop_process'):
            self.stop_process()
        else:
            # Fallback if stop_process was renamed or missing
            if hasattr(self, 'worker'):
                self.worker.stop()
            self.log("⏹ Process stopping...", "#DC3545")

    def start_sending(self, resume=False):
        if not self.service or not self.excel_path:
            QMessageBox.critical(self, "Error", "Authenticate and select Excel file first.")
            return
        
        items = self.list_drafts.selectedItems()
        if not items:
            QMessageBox.critical(self, "Error", "Select a draft.")
            return
        
        # self.drafts is Dict {FormattedText: ID}
        draft_key = items[0].text()
        draft_id = self.drafts.get(draft_key)
        
        if not draft_id:
             QMessageBox.critical(self, "Error", "Failed to identify draft ID.")
             return

        start_row = 2
        if resume:
            if os.path.exists(PROGRESS_FILE):
                with open(PROGRESS_FILE) as f:
                    start_row = json.load(f).get("last_row", 2)
            else:
                QMessageBox.information(self, "Info", "No progress file found. Starting from beginning.")

        # CC/BCC Inputs (Skip on Resume)
        if resume and hasattr(self, 'last_send_args'):
             # Use last used args if available
             cc_mode = self.last_send_args.get('cc_mode', 'none')
             cc_val = self.last_send_args.get('global_cc', '')
             bcc_mode = self.last_send_args.get('bcc_mode', 'none')
             bcc_val = self.last_send_args.get('global_bcc', '')
        elif resume:
             # Default if no history
             cc_mode, cc_val = 'none', ''
             bcc_mode, bcc_val = 'none', ''
        else:
             cc_mode, cc_val = self.ask_cc_bcc("CC")
             bcc_mode, bcc_val = self.ask_cc_bcc("BCC")
        
        # Store args for Real Start
        self.pending_send_args = {
             'service': self.service,
             'excel_path': self.excel_path,
             'draft_id': draft_id,
             'start_row': start_row,
             'cc_mode': cc_mode,
             'global_cc': cc_val,
             'bcc_mode': bcc_mode,
             'global_bcc': bcc_val,
             'display_name': self.display_name,
             'user_email': self.user_email,
             'total_rows': None # Will be updated if data loaded
        }
        
        # Add Attachment Rule
        if hasattr(self, 'attachment_empty_rule'):
             self.pending_send_args['attachment_empty_rule'] = self.attachment_empty_rule
        else:
             self.pending_send_args['attachment_empty_rule'] = 'yes'
             
        if resume:
             self.pending_send_args['is_resume'] = True
             self.real_start_sending()
        else:
             self.log("⏳ Loading preview data...", "#17A2B8")
             self.btn_process.setEnabled(False)
             self.btn_resume.setEnabled(False)
             
             self.data_loader = DataLoadingWorker(self.service, draft_id, self.excel_path)
             self.data_loader.data_loaded.connect(self.on_data_loaded)
             self.data_loader.status_signal.connect(lambda s: self.log(f"📋 {s}", "#17A2B8"))
             self.data_loader.error_signal.connect(self.on_error)
             self.data_loader.start()

    def on_data_loaded(self, draft_data, wb, all_headers, visible_headers, rows):
         # Open Dialog
         args = self.pending_send_args
         args['total_rows'] = len(rows) # Update total count
         
         dlg = AdvancedPreviewDialog(self, draft_data, all_headers, visible_headers, rows, 
                                     args['cc_mode'], args['global_cc'], 
                                     args['bcc_mode'], args['global_bcc'])
         
         dlg.start_sending.connect(self.real_start_sending)
         
         if dlg.exec_() != QDialog.Accepted:
             # Cancelled
             self.btn_process.setEnabled(True)
             if os.path.exists(PROGRESS_FILE): self.btn_resume.setEnabled(True)
             self.log("Preview cancelled workspace.", "#6C757D")

    def real_start_sending(self):
        args = self.pending_send_args
        
        # Disable UI
        self.btn_process.setEnabled(False)
        self.btn_resume.setEnabled(False)
        self.btn_stop.setEnabled(True)
        
        # Reset Progress if new start
        if args['start_row'] == 2:
            self.total_sent = 0
            self.total_failed = 0
            self.progress_bar.setValue(0) 
            self.apply_progress_style("#0d6efd")

        self.worker = EmailWorker(
            args['service'], args['excel_path'], args['draft_id'], args['start_row'], 
            args['cc_mode'], args['global_cc'], args['bcc_mode'], args['global_bcc'],
            args['display_name'], args['user_email'], args.get('total_rows'), args.get('is_resume', False),
            attachment_mode=self.chk_send_attachments.isChecked(),
            attachment_empty_rule=args.get('attachment_empty_rule', 'yes')
        )
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        # Removed preview_signal connection
        self.worker.live_preview_signal.connect(self.handle_live_preview_update)
        self.worker.stopped_signal.connect(self.on_stopped_stats)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error_signal.connect(self.on_error)
        self.worker.start()

    def on_stopped_stats(self, sent, failed, pending):
        # Accumulate totals
        self.total_sent += sent
        self.total_failed += failed
        
        msg_html = (
            f"<b>Session Stats:</b><br>"
            f"&nbsp;&nbsp;✅ Sent: <b>{sent}</b><br>"
            f"&nbsp;&nbsp;❌ Failed: <b>{failed}</b><br><br>"
            f"⏳ <b>Pending: {pending}</b>"
        )
        
        dlg = ModernInfoDialog(self, "Process Stopped", msg_html, "🛑", "#DC3545")
        dlg.exec_()

    def handle_live_preview_update(self, row_idx, row_values, status):
        # Show table if hidden
        if not self.table_preview.isVisible():
            self.table_preview.setVisible(True)
            self.table_preview.setFixedHeight(150) # Expand to show content

        # 1. Update Buffer
        # Remove existing entry for this row if present (to update status/position)
        self.live_preview_buffer = [item for item in self.live_preview_buffer if item[0] != row_idx]
        
        # Insert new/updated entry at TOP
        self.live_preview_buffer.insert(0, (row_idx, row_values, status))
        
        # 2. Render
        self.render_preview_table()

    def render_preview_table(self):
        self.table_preview.setRowCount(len(self.live_preview_buffer))
        
        for r_idx, (excel_row_idx, row_data, status) in enumerate(self.live_preview_buffer):
            # Filter data for columns
            filtered_data = [row_data[i] if i < len(row_data) else None for i in self.valid_header_indices]
            
            for c_idx, val in enumerate(filtered_data):
                item = QTableWidgetItem(str(val) if val is not None else "")
                item.setToolTip(str(val) if val is not None else "")
                self.table_preview.setItem(r_idx, c_idx, item)

            # Update Status Column Visuals
            status_col_idx = self.preview_header_map.get("Status")
            if status_col_idx is not None:
                # If status is passed explicitly, use it. Otherwise try to find it in data (for initial load)
                display_status = status
                if not display_status and len(row_data) > self.valid_header_indices[status_col_idx]:
                     # Fallback to data if no explicit status (e.g. initial load)
                     display_status = str(row_data[self.valid_header_indices[status_col_idx]])

                item = QTableWidgetItem(display_status)
                item.setTextAlignment(Qt.AlignCenter)
                
                if display_status == "Sent":
                    item.setBackground(QColor("#D4EDDA")) # Light Green
                    item.setForeground(QColor("#155724"))
                elif "Error" in display_status:
                    item.setBackground(QColor("#F8D7DA")) # Light Red
                    item.setForeground(QColor("#721C24"))
                elif display_status == "Sending...":
                    item.setBackground(QColor("#FFF3CD")) # Light Yellow
                    item.setForeground(QColor("#856404"))
                
                self.table_preview.setItem(r_idx, status_col_idx, item)
        
        # Hybrid Resizing Logic
        header = self.table_preview.horizontalHeader()
        
        # 1. Reset to Interactive (allows manual resize & keeps setColumnWidth)
        header.setSectionResizeMode(QHeaderView.Interactive)
        
        # 2. Auto-fit to content
        self.table_preview.resizeColumnsToContents()
        
        # 3. Clamp Max Width to 300px
        for i in range(self.table_preview.columnCount()):
            if self.table_preview.columnWidth(i) > 300:
                self.table_preview.setColumnWidth(i, 300)
        
        # 4. Check for Stretch (Fill empty space if table is small)
        total_width = sum(self.table_preview.columnWidth(i) for i in range(self.table_preview.columnCount()))
        viewport_width = self.table_preview.viewport().width()
        
        if total_width < viewport_width:
            # Distribute extra space manually to keep Interactive mode (allows resizing/scrolling)
            count = self.table_preview.columnCount()
            if count > 0:
                extra = viewport_width - total_width
                add_per_col = int(extra / count)
                for i in range(count):
                    self.table_preview.setColumnWidth(i, self.table_preview.columnWidth(i) + add_per_col)
        
        # Ensure scrollbar is always enabled if needed
        self.table_preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

    def ask_cc_bcc(self, type_name):
        mode = "individual"
        val = ""
        # We use standard dialogs for simplicity, but could be themed.
        # For now, just using standard PyQt functionality.
        reply = QMessageBox.question(self, f"{type_name} Options", 
                                     f"Use Global {type_name} for all emails?\n(No = use Excel column)", 
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            # Use custom resizable dialog
            dlg = ResizableInputDialog(f"Global {type_name}", f"Enter {type_name} email(s):", self)
            if dlg.exec_() == QDialog.Accepted:
                text = dlg.get_text()
                if text:
                    mode = "global"
                    val = text
            else:
                mode = "individual" # Fallback to individual if cancelled
        else:
            mode = "individual" # No = use Excel column
        return mode, val




    def on_toggle_attachments(self):
        if not self.chk_send_attachments.isChecked():
            # USER WANTS CONDITIONAL MODE
            # 1. Validation: Check if Excel has the column first!
            if hasattr(self, 'excel_path') and self.excel_path:
                try:
                    wb = openpyxl.load_workbook(self.excel_path, read_only=True)
                    ws = wb.active
                    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    headers = [str(c).strip().lower() for c in row1 if c] if row1 else []
                    wb.close()
                    
                    found = False
                    for name in ['send attachments', 'send attachment', 'attachment', 'attachments']:
                        if name in headers:
                            found = True
                            break
                    
                    if not found:
                        # ERROR: Column Missing
                        ModernInfoDialog(self, "Missing Column", 
                                       "<b>Cannot Enable Conditional Mode</b><br><br>"
                                       "Your Excel file is missing a <b>'Send Attachments'</b> column.<br>"
                                       "Please add this header to your Excel file and fill it with 'Yes' or 'No'.",
                                       "❌", "#DC3545").exec_()
                        
                        # Revert Checkbox
                        self.chk_send_attachments.blockSignals(True)
                        self.chk_send_attachments.setChecked(True)
                        self.chk_send_attachments.blockSignals(False)
                        return # STOP HERE
                        
                except Exception as e:
                    self.log(f"⚠️ Validation Error: {e}", "#FFC107")

            # OPEN NEW DIALOG only if validation passed
            dlg = ModernAttachmentDialog(self)
            res = dlg.exec_()
            
            if res == QDialog.Accepted:
                # Store the rule (yes/no)
                self.attachment_empty_rule = dlg.result_rule
                self.log(f"🔧 Conditional Attachments: Enabled (Empty Default: {self.attachment_empty_rule.upper()})", "#17A2B8")
            else:
                # Cancelled
                self.chk_send_attachments.blockSignals(True)
                self.chk_send_attachments.setChecked(True) # Revert
                self.chk_send_attachments.blockSignals(False)
        else:
            self.attachment_empty_rule = "yes" # Reset default

    def scan_excel_for_empty_attachments(self):
        # Helper to find empty rows for validtion
        if not hasattr(self, 'excel_path') or not self.excel_path: return []
        
        empty_rows = []
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() for c in ws[1]]
            
            # Identify columns
            col_email = -1
            col_name = -1
            col_att = -1
            
            for i, h in enumerate(headers):
                if h == 'email': col_email = i
                if h == 'name': col_name = i
                if h in ['attachment', 'attachments', 'send attachments', 'send attachment']: col_att = i
            
            if col_att != -1:
                # Scan
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Check Attachment Value
                    val = row[col_att] if len(row) > col_att else None
                    if not val or not str(val).strip():
                        # Found Empty
                        name = row[col_name] if col_name != -1 and len(row) > col_name else None
                        email = row[col_email] if col_email != -1 and len(row) > col_email else None
                        
                        # Strict Filter: Only show if Email is present (meaning it's a valid recipient)
                        if email and str(email).strip():
                             empty_rows.append([name if name else "(No Name)", email])
                        
                        if len(empty_rows) > 50: break # Limit
            
        except: pass
        return empty_rows

    def style_standard_button(self, btn, base_color_rgb):
        r, g, b = base_color_rgb
        btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: rgb({r}, {g}, {b});
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: rgb({max(0, r-30)}, {max(0, g-30)}, {max(0, b-30)});
            }}
            QPushButton:pressed {{
                background-color: rgb({max(0, r-60)}, {max(0, g-60)}, {max(0, b-60)});
            }}
            QPushButton:disabled {{
                background-color: #E9ECEF;
                color: #ADB5BD;
            }}
        """)

    def stop_process(self):
        if self.worker:
            self.worker.stop()
            self.btn_stop.setEnabled(False) # Immediately disable stop button
            self.log("🛑 Stop requested...", "#FFA500")
            
            # Change Color to RED to indicate stopped/failed (but keep progress value)
            self.apply_progress_style("#DC3545")

    def on_error(self, error_msg):
        self.log(f"CRITICAL ERROR: {error_msg}", "#DC3545")
        # Ensure UI reflects failure
        self.apply_progress_style("#DC3545")
        self.on_finished(-1, -1) # Trigger cleanup

    def on_finished(self, sent, failed):
        self.btn_process.setEnabled(True)
        # Enable Resume ONLY if a progress file exists (meaning it was stopped/interrupted)
        # Also, explicit check: if sent != -1 (completed), disable resume.
        if sent == -1 and os.path.exists(PROGRESS_FILE):
             self.btn_resume.setEnabled(True)
        else:
             self.btn_resume.setEnabled(False)
            
        self.btn_stop.setEnabled(False)
        self.worker = None # Cleanup worker reference
        
        # Change Color to GREEN if finished successfully (sent != -1)
        if sent != -1:
             # Accumulate final totals
             prev_sent = self.total_sent
             prev_failed = self.total_failed
             self.total_sent += sent
             self.total_failed += failed
             
             self.apply_progress_style("#198754")
             self.progress_bar.setValue(100) # Ensure it looks complete
             
             self.progress_bar.setValue(100) # Ensure it looks complete
             
             dlg = ModernResultDialog(self, prev_sent, prev_failed, sent, failed, self.total_sent, self.total_failed)
             dlg.exec_()
        else:
            # If stopped or failed, ensure it stays Red (but keep progress value)
            self.apply_progress_style("#DC3545")

if __name__ == "__main__":
    try:
        lock_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        lock_socket.bind(("localhost", 60105))
    except socket.error:
        print("App already running")
        sys.exit()

    app = QApplication(sys.argv)
    
    # Fonts
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    window = MailMergeApp()
    window.show()
    sys.exit(app.exec_())
